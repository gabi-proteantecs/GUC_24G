import os
import time

import pandas as pd
import pyvisa
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


class D2D_Subprogram:
    def __init__(self, gui):
        self.gui = gui
        self.i2c = None

    def DUT_32Bit_Load_cfl(self, slave, RegisterFile_path, excel_sheet):
        RW_row_datalog = []
        # for i in range(1):
        #     work_book = xlrd.open_workbook(RegisterFile_path)
        #     sheet = work_book.sheet_by_name(excel_sheet[i])
        #     sheet_row_Num = len(sheet.col(0))
        #     # print(All_sheet_list[i])
        #     for j in range(sheet_row_Num):
        #         row_datalog = sheet.row(j)
        #         if 'RW' in str(list(row_datalog)) :
        #             if str(list(row_datalog)).count('0x') >= 2:
        #                 name = str(list(row_datalog)[1]).split(':')[1].strip('\'') # cell to string
        #                 offset = str(list(row_datalog)[2]).split(':')[1].strip('\'').replace('0x', '') # cell to string
        #                 Startbit = int(float(str(list(row_datalog)[4]).split(':')[1]))  # cell to int
        #                 Endbit = int(float(str(list(row_datalog)[3]).split(':')[1]))  # cell to int
        #                 Bitlength = (Endbit-Startbit)+1
        #                 value = str(list(row_datalog)[6]).split(':')[1].strip('\'').replace('0x', '')  # cell to int
        #                 Delay = str(20)
        #                 note = str(list(row_datalog)[7]).split(':')[1].strip('\'')  # cell to int
        #                 row_datalog = (';'+ name+'\n'+
        #                                slave+' ', ' '+offset+' ',' '+str(Startbit)+' ',
        #                                ' '+(str(Bitlength))+' ',' '+value+' ',
        #                                ' '+Delay+'     ; '+note)
        #                 row_datalog = str(",".join(row_datalog))
        #                 RW_row_datalog += [row_datalog]
        #                 # print(RW_row_datalog)
        #             else:
        #                 pass
        #         else:
        #             pass
        # print(RW_row_datalog)
        # f = open(os.getcwd() + '/Test Report Log(EY0012A)/Register_Fille.def', 'w+')
        # info =  '; ==============================================================================\n'+\
        #         '; Description: \n'+\
        #         ';	Each register to be written must obey the following format ....\n'+\
        #         ';	slave, offset, startbit, bitlength, value, sleep time \n'+\
        #         '; \n'+\
        #         ';	slave: 7 bit I2C slave address (in Hex) 32 Bit\n'+\
        #         ';	offset: 4 bytes alignment (in Hex) 32 Bit\n'+\
        #         ';	startbit: 0~31 \n'+\
        #         ';	bitlength: 1~32 \n'+\
        #         ';	value: new value to be written (in Hex) 32 Bit\n'+\
        #         ';	sleep time: sleep time after writing this register. (in milliseconds) \n'+\
        #         '; ==============================================================================\n;'
        # f.write((info + '\n;\n'))
        # for i in range(len(RW_row_datalog)):
        #     row_data = RW_row_datalog[i]
        #     print(row_data)
        #     f.write((row_data+'\n'))
        # f.close()

    def RST(self, **kargs):
        visa = kargs.get("visa", "NA")

        rm = pyvisa.ResourceManager()
        E36000A = rm.open_resource(visa)
        print(E36000A.write("*RST"))

    def IDN(self, **kargs):
        visa = kargs.get("visa", "NA")

        try:
            rm = pyvisa.ResourceManager()
            model = rm.open_resource(visa)
            value = model.query("*IDN?")
        except Exception as e:
            value = "'visa command *IDN? error'"
            pass
            # print(f'\033{e}')
            # value = 'visa command *IDN? error'
            # print(f'\033{value}')

        return value

    def E363xA_Out_ON(self, **kargs):
        # print('Remote Keysight E363xA',flush=True)
        visa = self.gui.E363xA_visa.Value
        rm = pyvisa.ResourceManager()
        E363xA = rm.open_resource(visa)
        E363xA.write(":INST:NSEL 1;:OUTP 1;")

    def E363xA_Out_OFF(self, **kargs):
        # print('Remote Keysight E363xA',flush=True)
        visa = self.gui.E363xA_visa.Value
        rm = pyvisa.ResourceManager()
        E363xA = rm.open_resource(visa)
        E363xA.write(":INST:NSEL 1;:OUTP 0;")

    def E363xA_Setup(self, CH1_V, **kargs):
        # print('Remote Keysight E363xA',flush=True)
        # print('Power Supply Channe1 = ',CH1_V,'V\n', flush=True)
        visa = self.gui.E363xA_visa.Value
        rm = pyvisa.ResourceManager()
        E363xA = rm.open_resource(visa)
        E363xA.write(":INST:NSEL 1;:VOLT " + str(CH1_V) + ";:OUTP 1;")

    def E363xA_read_ch1_V(self):
        # print('Remote Keysight E363xA',flush=True)
        visa = self.gui.E363xA_visa.Value
        rm = pyvisa.ResourceManager()
        E363xA = rm.open_resource(visa)
        result = E363xA.query(":INST:NSEL 1;:MEAS:VOLT?")
        return result

    def E363xA_read_ch1_A(self):
        # print('Remote Keysight E363xA',flush=True)
        visa = self.gui.E363xA_visa.Value
        rm = pyvisa.ResourceManager()
        E363xA = rm.open_resource(visa)
        result = float(E363xA.query(":INST:NSEL 1;:MEAS:Curr?"))
        return f"E363xA_Channel1_Current = {result}"

    def E363xA_read_ch2_V(self):
        # print('Remote Keysight E363xA',flush=True)
        visa = self.gui.E363xA_visa.Value
        rm = pyvisa.ResourceManager()
        E363xA = rm.open_resource(visa)
        result = E363xA.query(":INST:NSEL 2;:MEAS:VOLT?")
        return result

    def E363xA_read_ch2_A(self):
        # print('Remote Keysight E363xA',flush=True)
        visa = self.gui.E363xA_visa.Value
        rm = pyvisa.ResourceManager()
        E363xA = rm.open_resource(visa)
        result = float(E363xA.query(":INST:NSEL 2;:MEAS:Curr?"))
        return f"E363xA_Channel2_Current = {result}"

    def E3631xA_Out_ON(self):
        visa = self.gui.E36300A_visa.Value
        # print('Remote Keysight E3631xA',flush=True)
        rm = pyvisa.ResourceManager()
        E3631xA_ = rm.open_resource(visa)
        for i in range(3):
            rm = pyvisa.ResourceManager()
            E3631xA = rm.open_resource(visa)
            E3631xA.write(":OUTP ON,(@" + str(i + 1) + ")")
            time.sleep(0.5)
            Output_Status = E3631xA.query("STAT:QUES:INST:ISUM" + str(i + 1) + ":COND?")
            # print(Output_Status)
            if str(Output_Status) == "+1\n":
                E3631xA.write(":OUTP OFF,(@" + str(i + 1) + ")")
                print(
                    "Error : Channel"
                    + str(i + 1)
                    + " is in CC (constant current) operating mode"
                )
            elif str(Output_Status) == "+3\n":
                E3631xA.write(":OUTP OFF,(@" + str(i + 1) + ")")
                print("Error : Channel" + str(i + 1) + " output has a hardware failure")
            else:
                pass

    def E3631xA_Out_OFF(self, **kargs):
        # print('Remote Keysight E3631xA',flush=True)
        visa = self.gui.E36300A_visa.Value

        rm = pyvisa.ResourceManager()
        E3631xA = rm.open_resource(visa)
        E3631xA.write("OUTPut:COUPle:CHANNel ALL")

    def E3631xA_Setup(self, CH1_V, CH2_V, CH3_V, CH1_A, CH2_A, CH3_A, **kargs):
        # print('Remote Keysight E3631xA',flush=True)
        # print(  'Power Supply Channe1 = ',CH1_V,'V\n'
        #         'Power Supply Channe2 = ',CH2_V,'V\n'
        #         'Power Supply Channe3 = ',CH3_V, 'V', flush=True)
        visa = self.gui.E36300A_visa.Value
        # Power supply voltage level Limit
        V_list = CH1_V, CH2_V, CH3_V
        A_list = CH1_A, CH2_A, CH3_A
        for i in range(3):
            rm = pyvisa.ResourceManager()
            E3631xA = rm.open_resource(visa)
            E3631xA.write(":OUTP ON,(@1,2,3)")
            E3631xA.write(
                "VOLT:MODE FIX,(@"
                + str(i + 1)
                + ");:VOLT "
                + str(V_list[i])
                + ",(@"
                + str(i + 1)
                + ");"
            )
            E3631xA.write(
                "CURR:MODE FIX,(@"
                + str(i + 1)
                + ");:CURR "
                + str(A_list[i])
                + ",(@"
                + str(i + 1)
                + ");"
            )
            E3631xA.write(":OUTP ON,(@" + str(i + 1) + ")")
            time.sleep(0.5)
            Output_Status = E3631xA.query("STAT:QUES:INST:ISUM" + str(i + 1) + ":COND?")
            # print(Output_Status)
            if str(Output_Status) == "+1\n":
                E3631xA.write(":OUTP OFF,(@" + str(i + 1) + ")")
                print(
                    "Error : Channel"
                    + str(i + 1)
                    + " is in CC (constant current) operating mode"
                )
            elif str(Output_Status) == "+3\n":
                E3631xA.write(":OUTP OFF,(@" + str(i + 1) + ")")
                print("Error : Channel" + str(i + 1) + " output has a hardware failure")
            else:
                pass
            """ 
            Output_Status info :
            Output_Status = 0 : The output is off or unregulated
            Output_Status = 1 : The output is in CC (constant current) operating mode
            Output_Status = 2 : The output is in CV (constant voltage) operating mode.
            Output_Status = 3 : The output has a hardware failure.
            """

    def E3631xA_Read_A(self, **kargs):
        # print('Remote Keysight E36233A',flush=True)
        visa = self.gui.E36300A_visa.Value

        VOLT_val_log = []
        CURR_val_log = []
        for i in range(3):
            rm = pyvisa.ResourceManager()
            E3631xA = rm.open_resource(visa)
            VOLT_val = E3631xA.query(":MEAS:VOLT? (@" + str(i + 1) + ")")
            CURR_val = E3631xA.query(":MEAS:CURR? (@" + str(i + 1) + ")")
            VOLT_val_log += [float(VOLT_val)]
            CURR_val_log += [float(CURR_val)]
        CH1_VOLT = VOLT = str(VOLT_val_log[0])
        CH2_VOLT = VOLT = str(VOLT_val_log[1])
        CH3_VOLT = VOLT = str(VOLT_val_log[2])
        CH1_CURR = CURR = str(CURR_val_log[0])
        CH2_CURR = CURR = str(CURR_val_log[1])
        CH3_CURR = CURR = str(CURR_val_log[2])
        print("Kesyight E3631xA Current Measure")
        return (
            f"CH1_VOLT={CH1_VOLT}"
            + ","
            + f"CH2_VOLT={CH2_VOLT}"
            + ","
            + f"CH3_VOLT={CH3_VOLT}"
            + ","
            + f"CH1_CURR={CH1_CURR}"
            + ","
            + f"CH2_CURR={CH2_CURR}"
            + ","
            + f"CH3_CURR={CH3_CURR}"
        )

    def E36233A_Out_OFF(self, **kargs):
        # print('Remote Keysight E36233A',flush=True)
        # print('Power Supply Output OFF', flush=True)
        visa = self.gui.E36233A_visa.Value

        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)
        E36233A.write("OUTPut:COUPle:CHANNel ALL")

    def E36233A_Out_ON(self, **kargs):
        visa = self.gui.E36233A_visa.Value

        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)
        E36233A.write(":OUTPut:STATe 1,(@1,2)")

    def E36233A_Out_OFF_RST(self, **kargs):
        # print('Remote Keysight E36233A',flush=True)
        # print('Power Supply Output OFF', flush=True)
        visa = self.gui.power_cycle_visa.Value

        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)
        E36233A.write("OUTPut:COUPle:CHANNel ALL")

    def E36233A_Out_ON_RST(self, **kargs):
        visa = self.gui.power_cycle_visa.Value

        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)
        E36233A.write(":OUTPut:STATe 1,(@1,2)")

    def E36233A_Out_OFF_RST_YQ(self, **kargs):
        visa = kargs.get("visa", "")

        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)
        E36233A.write("OUTPut:COUPle:CHANNel ALL")

    def E36233A_Out_ON_RST_YQ(self, **kargs):
        visa = kargs.get("visa", "")

        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)
        E36233A.write(":OUTPut:STATe 1,(@1,2)")

    def E36233A_setup(self, CH1_V, CH2_V, CH1_A, CH2_A, **kargs):
        sel_visa = kargs.get("sel_visa", 1)
        visa = kargs.get("visa", "")
        if sel_visa == 1:
            visa = self.gui.power_cycle_visa.Value
        V_list = CH1_V, CH2_V
        A_list = CH1_A, CH2_A
        for i in range(2):
            rm = pyvisa.ResourceManager()
            E36233A = rm.open_resource(visa)

            E36233A.write(
                "VOLT:MODE FIX,(@"
                + str(i + 1)
                + ");:VOLT "
                + str(V_list[i])
                + ",(@"
                + str(i + 1)
                + ");"
            )
            E36233A.write(
                "CURR:MODE FIX,(@"
                + str(i + 1)
                + ");:CURR "
                + str(A_list[i])
                + ",(@"
                + str(i + 1)
                + ");"
            )
            E36233A.write(":OUTP ON,(@" + str(i + 1) + ")")
            time.sleep(0.5)
            Output_Status = E36233A.query("STAT:QUES:INST:ISUM" + str(i + 1) + ":COND?")
            # print(Output_Status)
            if str(Output_Status) == "+1\n":
                E36233A.write(":OUTP OFF,(@" + str(i + 1) + ")")
                print(
                    "Error : Channel"
                    + str(i + 1)
                    + " is in CC (constant current) operating mode"
                )
            elif str(Output_Status) == "+3\n":
                E36233A.write(":OUTP OFF,(@" + str(i + 1) + ")")
                print("Error : Channel" + str(i + 1) + " output has a hardware failure")
            else:
                pass
            """ 
            Output_Status info :
            Output_Status = 0 : The output is off or unregulated
            Output_Status = 1 : The output is in CC (constant current) operating mode
            Output_Status = 2 : The output is in CV (constant voltage) operating mode.
            Output_Status = 3 : The output has a hardware failure.
            """

    def E36233A_setup_one_channel(self, CH_V, CH_A, CH_Select, **kargs):
        visa = kargs.get("visa", "")

        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)
        E36233A.write(
            "VOLT:MODE FIX,(@"
            + str(CH_Select)
            + ");:VOLT "
            + str(CH_V)
            + ",(@"
            + str(CH_Select)
            + ");"
        )
        E36233A.write(
            "CURR:MODE FIX,(@"
            + str(CH_Select)
            + ");:CURR "
            + str(CH_A)
            + ",(@"
            + str(CH_Select)
            + ");"
        )
        E36233A.write(":OUTP ON,(@" + str(CH_Select) + ")")
        Output_Status = E36233A.query("STAT:QUES:INST:ISUM" + str(CH_Select) + ":COND?")
        # print(Output_Status)
        if str(Output_Status) == "+1\n":
            E36233A.write(":OUTP OFF,(@" + str(CH_Select) + ")")
            print(
                "Error : Channel"
                + str(CH_Select)
                + " is in CC (constant current) operating mode"
            )
        elif str(Output_Status) == "+3\n":
            E36233A.write(":OUTP OFF,(@" + str(CH_Select) + ")")
            print("Error : Channel" + str(CH_Select) + " output has a hardware failure")
        else:
            pass
        """ 
        Output_Status info :
        Output_Status = 0 : The output is off or unregulated
        Output_Status = 1 : The output is in CC (constant current) operating mode
        Output_Status = 2 : The output is in CV (constant voltage) operating mode.
        Output_Status = 3 : The output has a hardware failure.
        """

    def E36233A_Read_A(self, **kargs):
        # print('Remote Keysight E36233A',flush=True)
        sep_curr = kargs.get("sep_curr", 1)
        visa = kargs.get("visa", "")

        CURR_val_log = []
        for i in range(2):
            rm = pyvisa.ResourceManager()
            E36233A = rm.open_resource(visa)
            CURR_val = E36233A.query(":MEAS:CURR? (@" + str(i + 1) + ")")
            CURR_val_log += [float(CURR_val)]
        CH1_CURR = str(CURR_val_log[0])
        CH2_CURR = str(CURR_val_log[1])
        # return (F'E36233A_Channel1 Current = {CH1_CURR} / E36233A_Channel2 Current =  {CH2_CURR}')
        Total_CURR = float(CURR_val_log[0]) + float(CURR_val_log[1])
        if sep_curr == 1:
            return CH1_CURR + "," + CH2_CURR
        else:
            return str(Total_CURR)

    def E36233A_Read_V(self, **kargs):
        # print('Remote Keysight E36233A',flush=True)
        visa = "USB0::0x2A8D::0x3302::MY59001241::0::INSTR"

        CURR_val_log = []
        for i in range(2):
            rm = pyvisa.ResourceManager()
            E36233A = rm.open_resource(visa)
            CURR_val = E36233A.query(":MEAS:VOLT? (@" + str(i + 1) + ")")
            CURR_val_log += [float(CURR_val)]
        CH1_CURR = str(CURR_val_log[0])
        CH2_CURR = str(CURR_val_log[1])
        # print(F'Channel1 Current = {CH1_CURR} /Channel2 Current =  {CH2_CURR}')
        # print(CH1_CURR+','+CH2_CURR)
        return CH1_CURR + "," + CH2_CURR

    def E3631xA_Read_V(self):
        # print('Remote Keysight E36233A',flush=True)
        visa = "USB0::0x2A8D::0x3302::MY59001241::0::INSTR"  # instument visa address

        CURR_val_log = []
        for i in range(3):  # read power supply 3 channel
            rm = pyvisa.ResourceManager()
            E3631xA = rm.open_resource(visa)
            CURR_val = E3631xA.query(
                ":MEAS:VOLT? (@" + str(i + 1) + ")"
            )  # Read meter voltage value command
            CURR_val_log += [float(CURR_val)]
        CH1_CURR = str(CURR_val_log[0])
        CH2_CURR = str(CURR_val_log[1])
        CH3_CURR = str(CURR_val_log[2])
        # print(CH1_CURR+','+CH2_CURR+','+CH3_CURR)
        return CH1_CURR + "," + CH2_CURR + "," + CH3_CURR

    def E3631xA_Read_V_ch1(self):
        # User can use this sub tools to control E3631xA keysight power supply
        import pyvisa  # import pyvisa

        rm = pyvisa.ResourceManager()  # call pyvisa function
        E3631xA = rm.open_resource(
            "USB0::0x2A8D::0x3302::MY59001241::0::INSTR"
        )  # instument visa address
        CURR_val = E3631xA.query(
            ":MEAS:VOLT? (@1"
        )  # Read power supply channel1 voltage value command

    def E36000A_setup_channel(self, **kargs):
        command = kargs.get("command", "1")  # instrument number on gui
        curr_limit = kargs.get("curr_limit", "max")

        buffer = command.split("/")
        INST_num = int(buffer[1])
        ch_select = buffer[2]
        ch_volt = buffer[3]
        ch_amp = buffer[4]

        if INST_num == 1:
            visa = self.gui.power1_visa_wx.Value
        if INST_num == 2:
            visa = self.gui.power2_visa_wx.Value
        if INST_num == 3:
            visa = self.gui.power3_visa_wx.Value
        if INST_num == 4:
            visa = self.gui.power4_visa_wx.Value
        if INST_num == 5:
            visa = self.gui.power5_visa_wx.Value
        rm = pyvisa.ResourceManager()
        E36233A = rm.open_resource(visa)

        E36233A.write(
            "VOLT:MODE FIX,(@"
            + ch_select
            + ");:VOLT "
            + ch_volt
            + ",(@"
            + ch_select
            + ");"
        )
        E36233A.write(
            "CURR:MODE FIX,(@"
            + ch_select
            + ");:CURR "
            + ch_amp
            + ",(@"
            + ch_select
            + ");"
        )
        E36233A.write(":OUTP ON,(@" + ch_select + ")")
        time.sleep(0.5)
        Output_Status = E36233A.query("STAT:QUES:INST:ISUM" + ch_select + ":COND?")
        # print(Output_Status)
        if str(Output_Status) == "+1\n":
            E36233A.write(":OUTP OFF,(@" + ch_select + ")")
            print(
                "Error : Channel"
                + ch_select
                + " is in CC (constant current) operating mode"
            )
        elif str(Output_Status) == "+3\n":
            E36233A.write(":OUTP OFF,(@" + ch_select + ")")
            print("Error : Channel" + ch_select + " output has a hardware failure")
        else:
            pass
        """ 
        Output_Status info :
        Output_Status = 0 : The output is off or unregulated
        Output_Status = 1 : The output is in CC (constant current) operating mode
        Output_Status = 2 : The output is in CV (constant voltage) operating mode.
        Output_Status = 3 : The output has a hardware failure.
        """

    def TA5000_Temp_Set(self, value):
        print("Set Test Temperature", value, "Degree", flush=True)
        rm = pyvisa.ResourceManager()
        Thermal_visa = self.gui.TA5000A_visa_wx.Value
        TA5000 = rm.open_resource(Thermal_visa)
        # print("SETP" + str(value))
        TA5000.write("SETP" + str(value))
        TA5000.write("Flow 1")

    def TA5000_Temp_off(self):
        Thermal_visa = self.gui.TA5000A_visa_wx.GetValue()
        rm = pyvisa.ResourceManager()
        TA5000 = rm.open_resource(Thermal_visa)
        TA5000.write("Flow 0")

    def TA5000_Temp_read(self, value, **kargs):
        sense = kargs.get("sense", 0)
        print("Remote MPI TA5000 Thermal", flush=True)
        rm = pyvisa.ResourceManager()
        Thermal_visa = self.gui.TA5000A_visa_wx.Value
        TA5000 = rm.open_resource(Thermal_visa)
        # TA5000 = rm.open_resource('TCPIP0::192.168.1.1::inst0::INSTR')
        Temp = float(TA5000.query("Temp?"))
        Deviation = 99
        while True:
            print("Temperature Now : " + str(Temp), flush=True)
            Temp = float(TA5000.query("Temp?"))
            if sense == 1:
                self.gui.run_0.Voltage_Sense(
                    no_rst=1, no_tx_rst=1, xls_ToolsEven_3=self.gui.ToolsEven_3
                )
                print(f"Voltage set without wait")
            else:
                time.sleep(1)
            Deviation = abs(value - Temp)
            if Deviation < 2:
                break
            else:
                pass
        TA5000.query("Temp?")

        return Temp

    def TA5000_Temp_read_temp(self, **kargs):
        rm = pyvisa.ResourceManager()
        Thermal_visa = self.gui.TA5000A_visa_wx.Value
        TA5000 = rm.open_resource(Thermal_visa)
        Temp = float(TA5000.query("Temp?"))
        return Temp

    def TA5000_Temp_read_only(self, **kargs):
        sense = kargs.get("sense", 0)
        rm = pyvisa.ResourceManager()
        Thermal_visa = self.gui.TA5000A_visa_wx.Value
        TA5000 = rm.open_resource(Thermal_visa)
        Temp = float(TA5000.query("Temp?"))
        print("Temperature Now : " + str(Temp), flush=True)
        if sense == 1:
            self.gui.run_0.Voltage_Sense(
                no_rst=1, no_tx_rst=1, xls_ToolsEven_3=self.gui.ToolsEven_3
            )

    def Key_34461A_Voltage(self, **kargs):
        visa = kargs.get("visa", "NA")

        rm = pyvisa.ResourceManager()
        M1_34461A = rm.open_resource(visa)
        M1_34461A_Voltage = float(M1_34461A.query(":SAMP:COUN 1;:TRIG:SOUR IMM;:READ?"))
        M1_34461A_Voltage = str(round(M1_34461A_Voltage, 3))
        return M1_34461A_Voltage

    def Key_34461A_Current(self, **kargs):
        visa = kargs.get("visa", "NA")

        rm = pyvisa.ResourceManager()
        model = rm.open_resource(visa)
        value = float(model.query(":SAMP:COUN 1;:TRIG:SOUR IMM;:READ?"))
        value = str(round(value / 0.001, 3))
        return value

    def KEI_DMM6500_Voltage(self, **kargs):
        # print('Remote KEITHLEY DMM6500')
        rm = pyvisa.ResourceManager()
        M5_DMM6500 = rm.open_resource(self.gui.meter3_visa_wx.Value)
        M5_DMM6500.write(":SENS:VOLTage:RANG:AUTO ON")
        M5_DMM6500_Voltage = float(M5_DMM6500.query(":READ? "))
        M5_DMM6500_Voltage = str(round(M5_DMM6500_Voltage, 3))
        return M5_DMM6500_Voltage

    def KEI_DMM6500_Temp(self, **kargs):
        visa = kargs.get("visa", "")

        rm = pyvisa.ResourceManager()
        M5_DMM6500 = rm.open_resource(visa)
        M5_DMM6500.write(":SENS:FUNC 'TEMP")
        M5_DMM6500.write(":SENS:TEMP:TC:TYPE K")
        M5_DMM6500_Voltage = float(M5_DMM6500.query(":READ? "))
        M5_DMM6500_Voltage = str(round(M5_DMM6500_Voltage, 3))
        return M5_DMM6500_Voltage

    def KEI_DMM6500_Voltage_YQ(self, **kargs):
        # print('Remote KEITHLEY DMM6500')
        rm = pyvisa.ResourceManager()
        M5_DMM6500 = rm.open_resource("USB0::0x05E6::0x6500::04516724::0::INSTR")
        M5_DMM6500.write(":SENS:VOLTage:RANG:AUTO ON")
        M5_DMM6500_Voltage = float(M5_DMM6500.query(":READ? "))
        M5_DMM6500_Voltage = str(round(M5_DMM6500_Voltage, 3))
        return M5_DMM6500_Voltage

    def KEI_DMM6500_Current(self, **kargs):
        visa = kargs.get("visa", "NA")
        name = kargs.get("name", "NA")

        rm = pyvisa.ResourceManager()
        model = rm.open_resource(visa)
        model.write(":SENS:FUNC 'CURRent';:SENS:CURRent:RANG:AUTO ON")
        value = float(model.query(":READ? "))
        value = str(round(value, 3))
        print(f"{name} curremt value={value}uA")
        return value

    def M3_34465A_Voltage(self, **kargs):
        # print('Remote Keysight 34465A')
        rm = pyvisa.ResourceManager()
        M3_34465A = rm.open_resource(self.gui.K_34465A_3_visa.Value)
        M3_34465A_Voltage = float(M3_34465A.query(":SAMP:COUN 1;:TRIG:SOUR IMM;:READ?"))
        M3_34465A_Voltage = str(round(M3_34465A_Voltage, 3))
        return M3_34465A_Voltage

    def M4_34411A_Voltage(self, **kargs):
        # print('Remote Keysight 34411A')
        rm = pyvisa.ResourceManager()
        M4_34411A = rm.open_resource(self.gui.K_34411A_4_visa.Value)
        M4_34411A_Voltage = float(M4_34411A.query(":SAMP:COUN 1;:TRIG:SOUR IMM;:READ?"))
        M4_34411A_Voltage = str(round(M4_34411A_Voltage, 3))
        return M4_34411A_Voltage

    def M5_34411A_Voltage(self, **kargs):
        visa = kargs.get("visa", "NA")

        rm = pyvisa.ResourceManager()
        M4_34411A = rm.open_resource(visa)
        M4_34411A_Voltage = float(M4_34411A.query(":SAMP:COUN 1;:TRIG:SOUR IMM;:READ?"))
        M4_34411A_Voltage = str(round(M4_34411A_Voltage, 3))
        return M4_34411A_Voltage

    def Meter_R_Current(self, **kargs):
        rm = pyvisa.ResourceManager()
        visa_list = (
            self.gui.K_34411A_4_visa.Value,
            self.gui.K_34461A_2_visa.Value,
            self.gui.K_34461A_1_visa.Value,
        )
        for i in range(3):
            visa = rm.open_resource(visa_list[i])
            Meter_R_Current = float(visa.query(":SAMP:COUN 1;:TRIG:SOUR IMM;:READ?"))
            Meter_R_Current = str(round(Meter_R_Current / 0.000001, 3))
            print(f"Meter_Channel{i + 1}={Meter_R_Current}uA")

    def Keysight_DataLog_793(self, **kargs):
        chan_array = kargs.get("chan_array", [])
        visa = kargs.get("visa", "")
        rm = pyvisa.ResourceManager()
        DataLog = rm.open_resource(visa)
        DataLog.write(":CONF:VOLT:DC (@101,102)")

        result = "NA"
        result_list = []
        chan_array = [
            "101",
            "102",
            "103",
            "104",
            "105",
            "106",
            "107",
            "108",
            "109",
            "110",
            "111",
            "112",
            "113",
            "114",
            "115",
            "116",
            "117",
            "118",
            "119",
            "120",
        ]
        # chan_array = ['101','102','103','104','105','106','107','108','109','110']
        for channel in chan_array:
            result = DataLog.query("MEAS:VOLT:DC? AUTO, DEF, (@" + channel + ")")
            result = round((float(result)), 4)
            result_list += [result]
            time.sleep(0.1)
        return result_list

    def Keysight_DataLog_793_101_104(self, **kargs):
        chan_array = kargs.get("chan_array", [])
        visa = kargs.get("visa", "")
        rm = pyvisa.ResourceManager()
        DataLog = rm.open_resource(visa)
        DataLog.write(":CONF:VOLT:DC (@101,102)")

        result = "NA"
        result_list = []
        chan_array = ["101", "102", "103", "104"]
        # chan_array = ['101','102','103','104','105','106','107','108','109','110']
        for channel in chan_array:
            result = DataLog.query("MEAS:VOLT:DC? AUTO, DEF, (@" + channel + ")")
            result = round((float(result)), 4)
            result_list += [result]
            time.sleep(0.1)
        return result_list

    def PG_81160A_2CH(self, **kargs):
        # print('Remote Keysight 81160A')
        CMU_S = kargs.get("CMU_S", 0)

        PG_81160A_visa = self.gui.pg1_visa_wx.Value
        rm = pyvisa.ResourceManager()
        PG_81160A = rm.open_resource(PG_81160A_visa)
        PG_81160A.write(
            ":FUNC1 SQU;:VOLT1 "
            + str(0.9)
            + ";:VOLT1:OFFS 0.000000;:FREQ1 "
            + str(50000000)
            + ";"
        )
        PG_81160A.write(
            ":FUNC2 SQU;:VOLT2 "
            + str(0.9)
            + ";:VOLT2:OFFS 0.000000;:FREQ2 "
            + str(CMU_S)
            + ";"
        )
        PG_81160A.write(
            ":OUTPut1:IMPedance:EXTernal 50;:OUTput1 ON;:OUTPut1:COMPlement ON;"
        )
        PG_81160A.write(
            ":OUTPut2:IMPedance:EXTernal 50;:OUTput2 ON;:OUTPut2:COMPlement ON;"
        )

    def S_LECROY_8254_AutoSetup(self, **kargs):
        # print('Remote LECROY 8254')
        visa = "USB0::0x05FF::0x1023::4205N20792::0::INSTR"
        rm = pyvisa.ResourceManager()
        LECROY_8254 = rm.open_resource(visa)
        LECROY_8254.write("ASET")

    def S_LECROY_8254_RCL(self, **kargs):
        # print('Remote LECROY 8254')
        visa = self.gui.Scope_visa.Value
        rm = pyvisa.ResourceManager()
        LECROY_8254 = rm.open_resource(visa)
        LECROY_8254.write("*RCL 1")

    def S_LECROY_8254_RCL3(self, **kargs):
        # print('Remote LECROY 8254')
        visa = "USB0::0x05FF::0x1023::4205N20792::0::INSTR"
        rm = pyvisa.ResourceManager()
        LECROY_8254 = rm.open_resource(visa)
        LECROY_8254.write("*RCL 3")

    def S_LECROY_8254_vol(self, **kargs):
        # print('Remote LECROY 8254')
        visa = self.gui.Scope_visa.Value
        rm = pyvisa.ResourceManager()
        LECROY_8254 = rm.open_resource(visa)
        LECROY_8254.write("CLSW")
        time.sleep(5)  # can't skip
        result = LECROY_8254.query("PAST? CUST, AVG")
        return result

    def S_LECROY_8254_vol_meas7(self, **kargs):
        # print('Remote LECROY 8254')
        visa = "USB0::0x05FF::0x1023::4205N20792::0::INSTR"
        rm = pyvisa.ResourceManager()
        LECROY_8254 = rm.open_resource(visa)
        LECROY_8254.write("CLSW")
        time.sleep(0.5)  # can't skip

        all_result = LECROY_8254.query("PAST? CUST, AVG")
        result = (((all_result.split(","))[9]).split("V"))[0]
        return result

    def xls_write_register(self, sheet_name, Test_Result):
        excel_path = "Test Report\Test Report.xlsx"

        Excel_Path = load_workbook(filename=excel_path)
        Select_sheet = Excel_Path[sheet_name]
        Start_row_Num = 4
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for a in range(len(Test_Result)):
            Result = (Test_Result[a]).split(",")
            for i in range(len(Result)):
                Select_sheet.cell(row=Start_row_Num + a, column=i + 1).alignment = align
                Select_sheet.cell(row=Start_row_Num + a, column=i + 1).border = border
                if "FAIL" in Test_Result:
                    font = Font(
                        "Arial",
                        size=10,
                        bold=True,
                        italic=False,
                        strike=False,
                        color="ff0000",
                    )
                else:
                    font = Font(
                        "Arial",
                        size=10,
                        bold=False,
                        italic=False,
                        strike=False,
                        color="000000",
                    )
                Select_sheet.cell(row=Start_row_Num + a, column=i + 1).font = font
                Select_sheet.cell(row=Start_row_Num + a, column=i + 1).value = Result[i]

        self.gui.graph.xls_save(Excel_Path, path=excel_path)
        Excel_Path.close()

    def xls_write_Hyperlink(self, **kargs):
        row = kargs.get("row", "NA")
        col = kargs.get("col", "NA")
        Hyperlink_path = kargs.get("Hyperlink_path", "NA")

        excel_path = "Test Report\Test Report.xlsx"
        sheet_name = "Test Result"

        Excel_Path = load_workbook(filename=excel_path)
        Select_sheet = Excel_Path[sheet_name]
        name = "Hyperlink_Eye"
        font = Font(
            "Arial", size=10, bold=False, italic=False, strike=False, color="0000ff"
        )
        HYPERLINK = '=HYPERLINK("' + Hyperlink_path + '",' + '"' + name + '"' + ")"
        Select_sheet.cell(row=row, column=col + 1).value = HYPERLINK
        Select_sheet.cell(row=row, column=col + 1).font = font

        self.gui.graph.xls_save(Excel_Path, path=excel_path)
        Excel_Path.close()

    def xls_write_result_vco(self, excel_path, sheet_name, Hyperlink_path, **kargs):
        Die0_4s_V = kargs.get("Die0_4s_V", 0)
        Die1_4s_V = kargs.get("Die1_4s_V", 0)
        Die2_4s_V = kargs.get("Die2_4s_V", 0)
        Die3_4s_V = kargs.get("Die3_4s_V", 0)
        Die0_8s_H = kargs.get("Die0_8s_H", 0)
        Die1_8s_H = kargs.get("Die1_8s_H", 0)
        Die2_8s_H = kargs.get("Die2_8s_H", 0)
        Die3_8s_H = kargs.get("Die3_8s_H", 0)
        excel_path = "Test Report\Test Report.xlsx"

        "close excel.exe"
        command = "taskkill /f /t /im EXCEL.exe"
        os.system(command)

        Excel_Path = load_workbook(filename=excel_path)
        Select_sheet = Excel_Path[sheet_name]
        for vco_result in [
            Die0_4s_V,
            Die1_4s_V,
            Die2_4s_V,
            Die3_4s_V,
            Die0_8s_H,
            Die1_8s_H,
            Die2_8s_H,
            Die3_8s_H,
        ]:
            for i in range(99999999999):
                Start_row = Select_sheet.cell(row=i + 5, column=1).value
                # print(Start_row)
                if Start_row == None:
                    break
                else:
                    pass
            Start_row_Num = i + 5

            # Start edit test value in excel report

            # print(Select_sheet, flush=True)
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

            align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for i in range(len(vco_result)):
                if vco_result[i] == "Hyperlink_Log":
                    font = Font(
                        "Arial",
                        size=10,
                        bold=False,
                        italic=False,
                        strike=False,
                        color="000000",
                    )
                    HYPERLINK = (
                        '=HYPERLINK("'
                        + Hyperlink_path
                        + '",'
                        + '"'
                        + vco_result[i]
                        + '"'
                        + ")"
                    )
                    Select_sheet.cell(row=Start_row_Num, column=i + 1).value = HYPERLINK
                    Select_sheet.cell(row=Start_row_Num, column=i + 1).font = font
                else:
                    font = Font(
                        "Arial",
                        size=10,
                        bold=False,
                        italic=False,
                        strike=False,
                        color="000000",
                    )
                    Select_sheet.cell(row=Start_row_Num, column=i + 1).font = font
                    Select_sheet.cell(
                        row=Start_row_Num, column=i + 1
                    ).value = vco_result[i]
                Select_sheet.cell(row=Start_row_Num, column=i + 1).alignment = align
                Select_sheet.cell(row=Start_row_Num, column=i + 1).border = border
        Excel_Path.save(excel_path)

    def xls_write_result_Others_VCO_EY0013A(self, **kargs):
        Test_Result = kargs.get("Test_Result", "Test Result")

        excel_path = "Test Report(EY0012A).xlsx"
        sheet_name = "Others_VCO_EY0013A"

        "close excel.exe"
        command = "taskkill /f /t /im EXCEL.exe"
        os.system(command)

        Excel_Path = load_workbook(filename=excel_path)
        Select_sheet = Excel_Path[sheet_name]
        for vco_result in [Test_Result]:
            for i in range(99999999999):
                Start_row = Select_sheet.cell(row=i + 5, column=2).value
                # print(Start_row)
                if Start_row == None:
                    break
                else:
                    pass
            Start_row_Num = i + 5

            # print(Select_sheet, flush=True)
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

            align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for i in range(len(vco_result)):
                font = Font(
                    "Arial",
                    size=10,
                    bold=False,
                    italic=False,
                    strike=False,
                    color="000000",
                )
                Select_sheet.cell(row=Start_row_Num, column=i + 2).font = font
                Select_sheet.cell(row=Start_row_Num, column=i + 2).value = vco_result[i]
                Select_sheet.cell(row=Start_row_Num, column=i + 2).alignment = align
                Select_sheet.cell(row=Start_row_Num, column=i + 2).border = border
        Excel_Path.save(excel_path)

    def xls_show_data(self, excel_path, sheet_name):
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        print(df)

    def DataLog_E3631xA_sense(self, **kargs):
        min_sense_v = kargs.get("min_sense_v", 0.4)
        print("DataLog_voltage_sense")
        Power_setup = (
            str(self.gui.E3631xA_CH1_V.Value)
            + ","
            + str(self.gui.E3631xA_CH2_V.Value)
            + ","
            + str(self.gui.E3631xA_CH3_V.Value)
        )
        Power_setup_list = Power_setup.split(",")
        CH1_V = Power_setup_list[0]
        CH2_V = Power_setup_list[1]
        CH3_V = Power_setup_list[2]
        CH1_A = self.gui.E3631xA_CH1_A_Limit.Value
        CH2_A = self.gui.E3631xA_CH2_A_Limit.Value
        CH3_A = self.gui.E3631xA_CH3_A_Limit.Value
        self.E3631xA_Setup(CH1_V, CH2_V, CH3_V, CH1_A, CH2_A, CH3_A)
        time.sleep(2)
        Voltage_sense = self.M6_34970A()  # Datalog voltage senses Channel1/2/3

        Power_setup_list = []
        S_CH_V = [0.0, 0.0, 0.0]
        for i in range(3):
            PowerSupply = float((Power_setup.split(","))[i])
            S_CH_V[i] = float((Voltage_sense.split(","))[i])
            Voltage_Gap = PowerSupply - S_CH_V[i]
            Power_val = PowerSupply + Voltage_Gap
            Power_setup_list += [str(Power_val)]

        if (
            S_CH_V[0] < min_sense_v
            or S_CH_V[1] < min_sense_v
            or S_CH_V[2] < min_sense_v
        ):
            print(
                f"\033Sense too small V1:{S_CH_V[0]}, V2:{S_CH_V[1]} or V3:{S_CH_V[2]} < {min_sense_v}"
            )
            raise BaseException("Error")
        else:
            CH1_V = Power_setup_list[0]
            CH2_V = Power_setup_list[1]
            CH3_V = Power_setup_list[2]
            self.E3631xA_Setup(CH1_V, CH2_V, CH3_V, CH1_A, CH2_A, CH3_A)
            self.E3631xA_Out_ON()
            print(f"VoltageSense 101~103, Get={Voltage_sense}, Set={Power_setup_list}")

    def DataLog_E3631xA_VDDC_104_sense(self, **kargs):
        ToolsEven_3 = kargs.get("ToolsEven_3", "Even-2:0.75")
        direct_set = kargs.get("direct_set", 0)
        direct_set_v = kargs.get("direct_set_v", 0.675)
        Even_str = ToolsEven_3
        Power_val = float((Even_str.split(":"))[1]) if direct_set == 0 else direct_set_v
        Curr_CH1 = self.gui.E36233A_CH1_A_Limit.Value
        Curr_CH2 = self.gui.E36233A_CH2_A_Limit.Value
        self.E36233A_setup(Power_val, Power_val, Curr_CH1, Curr_CH2)
        # time.sleep(w_time)

    def DPO72504C_Save_png(self, **kargs):
        File_name = kargs.get("File_name", "NA")
        visa = "GPIB0::1::INSTR"
        rm = pyvisa.ResourceManager()
        DPO72504C = rm.open_resource(visa)
        Save_png_command = (
            "EXPort:FILEName 'C:\\Test Waveform\\" + File_name + ".png';:EXPort STAR"
        )
        DPO72504C.write("DISplay:PERSistence:RESET")
        time.sleep(120)
        DPO72504C.write(":EXPORT: FORMAT PNG")
        DPO72504C.write(Save_png_command)

    def IT6300_Output_en(self):
        visa = "USB0::0xFFFF::0x6300::600068011717630047::0::INSTR"
        rm = pyvisa.ResourceManager()
        IT6300 = rm.open_resource(visa)
        IT6300.write("OUTPut:STATe:ALL 0")
        time.sleep(1)
        IT6300.write("OUTPut:STATe:ALL 1")
