import ctypes
import datetime
import os
import pickle
import re
import subprocess
import sys
import time
import tkinter as tk
import tkinter.font as tkFont
from collections import OrderedDict
from tkinter import *

import numpy as np
import openpyxl
import pandas
import pandas as pd
import psutil
import pyautogui
import wx  # D2D use
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont

# gen exe code need
import Glink_phy
import Glink_run
import gui  # import the newly created GUI file by wxformbuilder
from Instrument import D2D_Subprogram
from Raspberry_Pico import *


class RedirectText(object):
    def __init__(self, aWxRichTextCtrl):
        self.out = aWxRichTextCtrl

    def write(self, string):
        self.out.MoveEnd()
        if "\b" in string:
            self.out.MoveUp()
        if "\033" in string:
            self.out.BeginTextColour((255, 0, 0))
        elif "\034" in string:
            self.out.BeginTextColour((102, 255, 255))
        else:
            self.out.BeginTextColour((255, 255, 255))
        self.out.WriteText(string)
        self.out.ScrollPages(1)
        self.out.EndTextColour()

    def flush(self):
        try:
            wx.GetApp().Yield()
        except:
            pass


class MainFrame(gui.MainFrame):
    def __init__(self, parent):
        gui.MainFrame.__init__(self, parent)
        self.rst_visa = "USB0::0x2A8D::0x3302::MY59001241::0::INSTR"
        self.EHOST = [
            [0x01, 0x02, 0x03],  # Die0 tport/H/V
            [0x01, 0x02, 0x03],  # Die1 tport/H/V
            [0x01, 0x02, 0x03],
        ]  # Die3 tport/H/V
        self.eye_graph_en = 0
        self.test_str_org = ["Load"]
        self.pass_fail = "NA"
        self.Temp_now = "NA"
        self.i2c = None
        self.phy_0 = None
        self.Spec = None
        self.run_0 = None
        self.jtag = None
        self.GROUP = 4
        self.SLICE = 8
        self.vef_num = 64
        self.slice = [0, 1, 2, 3]
        self.Power_Select = ""
        self.visa = D2D_Subprogram(self)
        self.TP_use = 0
        self.abp_en = 0
        self.ini_wo_reset = 1
        self.font = "Arial"
        self.m_richText1.SetBackgroundColour(wx.Colour(16, 16, 16))
        self.m_richText1.SetDefaultStyle(wx.TextAttr(wx.WHITE, (16, 16, 16)))
        self.TestItem_Now_wx.SetBackgroundColour(wx.Colour(16, 16, 16))
        self.TestItem_Now_wx.SetDefaultStyle(wx.TextAttr(wx.WHITE, (16, 16, 16)))
        sys.stdout = RedirectText(self.m_richText1)
        self.info_window_wx.Selection = 2
        self.def_gui()
        self.tools_path = os.path.dirname(os.path.abspath(__file__))

        # Load pll map for gui tree use
        # dirPath = r"Test Report\Register Map\PLL_Map"
        # result = (next(os.walk(dirPath))[2])[0]
        # self.pll_map_path = f'{dirPath}/{result}'
        # load pll map sheet name
        # xl = pandas.ExcelFile(self.pll_map_path)
        # sheet = xl.sheet_names
        # self.pll_sheet_name_tree.SetItems(sheet)

        # Load slice map for gui tree use
        # dirPath = r"Test Report\Register Map\Slice_Map"
        # result = (next(os.walk(dirPath))[2])[0]
        # self.slice_map_path = f'{dirPath}/{result}'
        # load pll map sheet name
        # xl = pandas.ExcelFile(self.slice_map_path)
        # sheet = xl.sheet_names
        # self.slice_sheet_name_tree.SetItems(sheet)

        self.datalog1_visa_chk()
        self.datalog2_visa_chk()
        self.power_channel_set()
        # self.reg_tree_event()

    def connect(self, event):
        self.eye_scan_en = 0
        self.bypass_report = 0
        self.info_window_wx.Selection = 2
        self.eye_scan_cycle = 0
        self.Register_init_en = 1
        self.eye_scan = "1d"
        ip_list = ("EZ0005A", "EZA001A")
        ip_num = self.ip_version_wx.GetSelection()
        self.ip_version = ip_list[ip_num]
        ip_list = ("UCIe_2p5D", "GLink_3D")
        ip_num = self.ip_version_wx.GetSelection()
        self.spec_version = ip_list[ip_num]

        self.xls_report_path = "Test Report\\\Test Report EZ0005A.xlsx"

        self.info_window_wx.Selection = 1
        self.m_richText1.Clear()
        print("Start Connect Test Chip", flush=True)
        self.eye_graph_info_wx.SetBitmap(
            wx.Bitmap("TestTools/blank.png", wx.BITMAP_TYPE_ANY)
        )

        try:
            f = open("TestTools\\project.json", "r")
            json_data = f.read()
            self.project_json = json.loads(json_data)

            for p in self.project_json.keys():
                if p == "i2c_project_speed_mode_test" or "//" in p:
                    continue
                else:
                    pass

        except FileNotFoundError:
            print("file project.json does not exist")
            self.project_json = False

        print(f"json change project start")
        get_json = self.project_select(True)
        print(f"get_json check = {get_json}")

        if (self.m_richText1.GetValue()).find("COM") != -1:
            buffer = (((self.m_richText1.GetValue()).split("COM"))[1])[0]
            self.I2C_info.SetBackgroundColour(colour="green")
            self.I2C_info.Value = f"I2C Module Raspberry Pico Initialization Pass"
            # self.I2C_info.Value = f'GUC UCIe 32Gb/s Eye Diagram'
        else:
            self.I2C_info.SetBackgroundColour(colour="red")
            self.I2C_info.Value = f"I2C Module Raspberry Pico Initialization Fail"

        if (self.ip_version).find("Version") != -1:
            self.m_richText1.Clear()
            print(f"\033 Please Select IP Version !!")
        # self.phy_0.set_input_pin6()

    def run(self, event):
        # self.run_0.thermal_die_CHK()
        # aa = self.visa.IDN(visa='USB0::0x2A8D::0x3302::MY61001409::0::INSTR')
        self.sw_en = 0
        self.m_richText1.Clear()
        self.phase_num = 64
        self.vref_num = 32
        self.driving_strength_en = 0
        self.flow_control_en = 0
        self.vref_start = "0x00"
        self.info_window_wx.Selection = 1
        voltage_sense = self.voltage_sense.Value
        if voltage_sense == True:
            self.avdd_sense_en = 1
        else:
            self.avdd_sense_en = 0

        start_time = datetime.datetime.now()
        self.m_textCtrl9.Value = "Start Test"
        self.pass_fail = None

        try:
            if self.m_toggleBtn_run_test.GetValue() == 1:
                self.m_toggleBtn_run_test.Label = "Disable"

            self.m_richText1.Clear()

            self.Test_step_Now = 0
            for run_n in range(int(self.Test_Cycle_wx.Value)):
                self.m_richText1.Clear()
                self.eye_scan_en = 0

                self.Chip_Corner = self.Corner_Version_wx.Value
                self.chip_number = self.chip_number_wx.Value

                # AutoTest
                self.info_window_wx.Selection = 1
                self.Autotest_en = 1
                workbook = openpyxl.load_workbook(self.xls_report_path, data_only=True)
                sheet = workbook["AutoTest Condtion"]
                row_s = 7  # for AutoTest Codition sheet use
                temp = []
                onoff = []

                for i in range(sheet.max_row - 6):
                    temp_val = sheet.cell(row=i + row_s, column=3).value
                    onoff_val = sheet.cell(row=i + row_s, column=1).value
                    temp += [temp_val]
                    onoff += [onoff_val]
                self.NowTemp_list = list(OrderedDict.fromkeys(temp))
                self.auto_on_num = len(list(filter(lambda x: "ON" in x, onoff)))
                # Temperature flow sequence from excel file
                buffer = sheet.cell(row=3, column=3).value
                if (str(buffer)).find("/") != -1:
                    buffer = buffer.split("/")
                    self.ToolsTemp_list = list(map(float, buffer))
                else:
                    self.ToolsTemp_list = [buffer]
                for k in range(len(self.ToolsTemp_list)):
                    self.Temp_now = self.ToolsTemp_list[k]
                    Seach_temp = self.ToolsTemp_list.count(self.Temp_now)
                    # Test temperature Loop
                    if Seach_temp == 0:
                        pass
                    else:
                        if self.ThermalOn_OFF.Value:
                            # Termal Air CTL
                            self.visa.TA5000_Temp_Set(self.Temp_now)
                            time.sleep(1)
                            self.visa.TA5000_Temp_read(self.Temp_now)
                            Test_Log = (self.m_richText1.Value).strip()
                            time.sleep(1)
                            # Termal instrument delay
                            Temp_Delay = int(self.Termal_Delay.Value)
                            for i in range(Temp_Delay):
                                time.sleep(1)
                                # print(f'\bTemperature Delay :{Temp_Delay:>5}', flush=True, end='')
                                self.m_richText1.Clear()
                                print(Test_Log, "\n")
                                print("Test Temperature", (self.Temp_now, "Degree"))
                                print(
                                    "Temperature Delay :", +(Temp_Delay - i), flush=True
                                )
                        # Check AutoTest all test condition
                        # Row or column values must be at least 1
                        for R in range(sheet.max_row):
                            self.item_S_time = datetime.datetime.now()
                            self.Test_OnOff = sheet.cell(row=row_s + R, column=1).value
                            self.Item_Temp_value = sheet.cell(
                                row=row_s + R, column=3
                            ).value
                            if (
                                self.Test_OnOff == "ON"
                                and self.Temp_now == self.Item_Temp_value
                            ):
                                self.def_gui(clear_count=0)

                                self.m_richText1.Clear()
                                self.phy_0.set_input_pin6()

                                self.info_window_wx.Selection = 1
                                self.Test_step_Now = self.Test_step_Now + 1
                                print("< Auto Test Condition >")
                                Test_Fun_ID = sheet.cell(row=row_s + R, column=2).value
                                # Test function list
                                Fun_str = sheet.cell(row=3, column=2).value
                                Fun_list = Fun_str.split("\n")
                                self.TestItem_full = (Fun_list[Test_Fun_ID + 0])[3:]
                                self.TestItem = ((self.TestItem_full).split("("))[0]
                                print(f"Test Item : {self.TestItem}")
                                print("Test Temperature : ", self.Temp_now, "Degree")
                                self.chip_version = (
                                    str(self.Chip_Corner) + "_" + str(self.chip_number)
                                )
                                print(f"PVT Corner : {self.chip_version}")
                                self.TestDataRate = sheet.cell(
                                    row=row_s + R, column=4
                                ).value
                                print(f"Test DataRate : {self.TestDataRate}")

                                # for D2D/UCIe Test
                                if (self.TestItem).find("Specialized") == -1:
                                    self.mode_name = sheet.cell(
                                        row=row_s + R, column=5
                                    ).value
                                    self.Chip_Mode = self.mode_name
                                    print(f"Test Mode : {self.Chip_Mode}")
                                    worst_function = [
                                        "Seach Eye High Min",
                                        "Seach Eye Width Min",
                                        "Seach Phase Fail Max",
                                        "Merge All Slice",
                                    ]
                                    self.eye_scan_type.Items = [
                                        "Die2_S0",
                                        "Die2_S2",
                                        "Die0_S0",
                                        "Die0_S2",
                                    ] + worst_function  # by chip
                                    self.eye_scan_type.Selection = 0  # by chip
                                    if (
                                        sheet.cell(row=row_s + R, column=6).value
                                        == "ON"
                                    ):
                                        self.Training_set = "Enable"
                                    else:
                                        self.Training_set = "Disable"
                                    print(f"Hardware Training : {self.Training_set}")
                                    rst_list = (
                                        "Hardware Reset (GPIO_Reset)",
                                        "Hardware Reset (Power_Reset)",
                                        "Skip Reset",
                                    )
                                    self.sys_rst_num = sheet.cell(
                                        row=row_s + R, column=7
                                    ).value
                                    print(f"Chip Reset : {rst_list[self.sys_rst_num]}")
                                    self.BIST_loop = sheet.cell(
                                        row=row_s + R, column=8
                                    ).value
                                    self.BIST_time = sheet.cell(
                                        row=row_s + R, column=9
                                    ).value
                                    print(
                                        f"BIST Test Loop : {self.BIST_loop} / BIST Test Time : {self.BIST_time}"
                                    )
                                    self.map_col_name = sheet.cell(
                                        row=row_s + R, column=10
                                    ).value
                                    self.register_setup_name = sheet.cell(
                                        row=row_s + R, column=11
                                    ).value
                                    print(
                                        f"Excel Register Setup Sheet Seaunece Name : {self.register_setup_name}"
                                    )
                                    self.File_name = sheet.cell(
                                        row=row_s + R, column=12
                                    ).value
                                    if self.File_name == None:
                                        self.File_name = ""
                                        print(f"FileName : None")
                                    else:
                                        print(f"FileName : {self.File_name}")
                                    self.note = sheet.cell(
                                        row=row_s + R, column=13
                                    ).value
                                    print(f"Note : {self.note}")

                                    # self.visa.E36233A_Out_OFF_RST_YQ()
                                    # time.sleep(2)
                                    # self.visa.E36233A_Out_ON_RST_YQ()
                                    # time.sleep(2)
                                    self.GUC_chip_rst()

                                    all_power_en = sheet.cell(
                                        row=row_s + R, column=30
                                    ).value
                                    if (
                                        all_power_en != "NA"
                                        and all_power_en != 0
                                        and all_power_en != None
                                    ):
                                        # use all power banana connect
                                        power1 = sheet.cell(
                                            row=row_s + R, column=31
                                        ).value
                                        power2 = sheet.cell(
                                            row=row_s + R, column=32
                                        ).value
                                        self.visa.E36233A_setup(
                                            power1, power2, "20", "20"
                                        )
                                        self.visa.E36233A_Out_ON_RST()
                                        time.sleep(2)
                                    else:
                                        pass
                                    for num in range(10):
                                        command = (
                                            sheet.cell(
                                                row=row_s + R, column=14 + num
                                            ).value
                                        ).upper()
                                        if command == "BYPASS":
                                            pass
                                        elif command != None:
                                            self.power_source_autotest(
                                                command=command, num=num
                                            )
                                        else:
                                            pass
                                        if self.avdd_sense_en == 1:
                                            self.voltage_sense_avdd = (
                                                (
                                                    (
                                                        sheet.cell(
                                                            row=row_s + R, column=14 + 4
                                                        ).value
                                                    ).upper()
                                                ).split("/")
                                            )[1]
                                        else:
                                            self.voltage_sense_avdd = 0.75

                                    # test even list
                                    even_list = []
                                    for e in range(20):
                                        even = sheet.cell(
                                            row_s + R, column=34 + e
                                        ).value
                                        if even != "NA" and even != None:
                                            print(f"Test Even{e + 1} : {even}")
                                        else:
                                            even = "NA"
                                        even_list += [even]
                                    self.even_1 = self.eye1D2D = even_list[0]
                                    self.even_2 = even_list[1]
                                    if self.even_2 == 0:
                                        self.log_type = "NON"
                                    elif self.even_2 == 1:
                                        self.log_type = "RD"
                                    else:
                                        self.log_type = "NON"
                                    self.even_3 = self.sheet_num = even_list[2]
                                    if self.sheet_num == "NA":
                                        self.sheet_num = 1
                                    else:
                                        self.sheet_num = int(self.sheet_num)
                                    self.even_4 = even_list[
                                        3
                                    ]  # pcs bist check pi cycles
                                    self.even_5 = even_list[
                                        4
                                    ]  # scan pi satus , after HW training
                                    self.even_6 = even_list[5]  # pmad test pattern set
                                    self.even_7 = even_list[6]  # sw pi range set
                                    self.even_8 = self.lane_valid_en = even_list[
                                        7
                                    ]  # Lane "valid" enable
                                    self.even_9 = self.sw_bist = even_list[
                                        8
                                    ]  # software training test - test mode select
                                    self.even_10 = even_list[
                                        9
                                    ]  # software training test 1 lane
                                    self.even_11 = even_list[10]
                                    self.even_12 = even_list[11]
                                    self.even_13 = even_list[12]
                                    self.even_14 = even_list[13]
                                    self.even_15 = even_list[14]
                                    self.note = (
                                        f"{self.note},"
                                        f"{self.even_1},{self.even_2},{self.even_3}"
                                        f"{self.even_4},{self.even_5},{self.even_6}"
                                        f"{self.even_7},{self.even_8},{self.even_9}"
                                        f"{self.even_10}"
                                    )

                                    self.test_condition = []

                                    for s in range(51):
                                        buffer = sheet.cell(
                                            row_s + R, column=s + 3
                                        ).value
                                        self.test_condition += [buffer]

                                    self.abp_en = 1
                                    self.ChkLog_abp(find="error!! off_len=8")
                                    # buffer = self.visa.KEI_DMM6500_Voltage_YQ()
                                    # print(f'IOVDD Power Level={buffer}V', flush=True)
                                    self.Start_Test(run_n=run_n)
                                else:
                                    # Run Specialized Function
                                    self.m_richText1.Clear()
                                    print(f"Test Item : {self.TestItem_full}")
                                    print(
                                        "Test Temperature : ", self.Temp_now, "Degree"
                                    )
                                    self.chip_version = (
                                        str(self.Chip_Corner)
                                        + "_"
                                        + str(self.chip_number)
                                    )
                                    print(f"PVT Corner : {self.chip_version}")
                                    self.TestDataRate = sheet.cell(
                                        row=row_s + R, column=4
                                    ).value
                                    print(f"Test DataRate : {self.TestDataRate}")
                                    self.mode_name = sheet.cell(
                                        row=row_s + R, column=5
                                    ).value
                                    self.Chip_Mode = self.mode_name
                                    print(f"Test Mode : {self.Chip_Mode}")
                                    self.register_setup_name = sheet.cell(
                                        row=row_s + R, column=11
                                    ).value
                                    print(
                                        f"Excel Register Setup Sheet Seaunece Name : {self.register_setup_name}"
                                    )
                                    self.File_name = sheet.cell(
                                        row=row_s + R, column=12
                                    ).value
                                    if self.File_name == None:
                                        self.File_name = ""
                                        print(f"FileName : None")
                                    else:
                                        print(f"FileName : {self.File_name}")
                                    self.note = sheet.cell(
                                        row=row_s + R, column=13
                                    ).value
                                    print(f"Note : {self.note}")
                                    self.Start_Test_Specialized()
                        else:
                            pass

            print("\nElapsed ALL Item : ", datetime.datetime.now() - start_time)
        except Exception as e:
            print(e)
            print("--- Run Interrupt With Fail ---\n\n", flush=True)
            self.m_textCtrl9.Value = "Run Fail"
            self.m_toggleBtn_run_test.SetValue(False)
            self.m_toggleBtn_run_test.Label = "Run"

        Termal_OnOFF = self.ThermalOn_OFF.Value
        if Termal_OnOFF == False:
            pass
        else:
            self.visa.TA5000_Temp_off()

        self.phy_0.pico_gpio_low(25, 0)  # gpio number , 0:pull low 1:pull high
        print("--- TestTools Done ---\n\n", flush=True)
        # self.m_textCtrl9.Value = 'Test Done'
        self.m_toggleBtn_run_test.SetValue(False)
        self.m_toggleBtn_run_test.Label = "Run"

        if self.eye_graph_en == 1:
            buffer = 0
        else:
            buffer = 1
            self.eye_graph_info_wx.SetBitmap(
                wx.Bitmap("TestTools/blank.png", wx.BITMAP_TYPE_ANY)
            )
        self.info_window_wx.Selection = buffer

    def Register_init(self):
        xls = pd.ExcelFile("Test Report\\\Test Report EZ0005A.xls")

        # load PLL Register Setup in excel
        df = xls.parse(self.sheet_num)
        # PLL Map
        print(f"\034Load training/bist register value in excel", flush=True)

        reg_col_num = self.seach_xls_reg(df, "Register Name")
        if self.reg_sequence_loops - 1 == reg_col_num:
            print(
                f'\033The  " {self.register_setup_name} "  register sequence read failed'
            )
        else:
            self.pll_en_reg = self.pll_seach_xls_str(
                df, "PLL", "HW1", reg_vai_num=reg_col_num
            )
            self.hw_non_1 = self.pll_seach_xls_str(
                df, "HW1", "PLL2", reg_vai_num=reg_col_num
            )
            self.PLL_NLB = self.pll_seach_xls_str(
                df, "PLL2", "HW2", reg_vai_num=reg_col_num
            )
            self.hw_NLB = self.pll_seach_xls_str(
                df, "HW2", "PCS1", reg_vai_num=reg_col_num
            )
            self.PCS_BIST_Check_NON = self.pll_seach_xls_str(
                df, "PCS1", "PMAD1", reg_vai_num=reg_col_num
            )
            self.PMAD_BIST_Check_NON = self.pll_seach_xls_str(
                df, "PMAD1", "USER1", reg_vai_num=reg_col_num
            )
            self.User_Define_Squence_1 = self.pll_seach_xls_str(
                df, "USER1", "USER2", reg_vai_num=reg_col_num
            )
            self.User_Define_Squence_2 = self.pll_seach_xls_str(
                df, "USER2", "DONE", reg_vai_num=reg_col_num
            )

        # print(f'\034Load function test register value in excel', flush=True)
        # self.func.Register_init(xls_report_path=self.xls_report_path)

        # print(f'\034Load (PMAD) lane pattern register value in excel', flush=True)
        # df = xls.parse(3)
        # reg_col_num = self.seach_xls_reg_lane(df, 'Lane')
        # self.lane_seach_xls_str(df, reg_col_num=reg_col_num)

    def TestItem_Init(self, **kwargs):
        TestItem = kwargs.get("TestItem", "0")

        self.log_name()
        self.TestResult = [""]
        self.eye_scan_en = 0

        # run HW_Training only or single test
        if TestItem == "ABP_Enable":
            self.eye_graph_en = 0
            self.Register_init_en = 0
            self.phy_0.indirect_enable(0, 1)
            self.phy_0.indirect_enable(0, 2)
            self.phy_0.indirect_enable(1, 1)
            self.phy_0.indirect_enable(1, 2)
            self.phy_0.indirect_enable(2, 1)
            self.phy_0.indirect_enable(2, 2)
            self.ChkLog_fail(find="failed")
            self.Test_Info_list = [
                str(self.pass_fail),
                self.pll_LOL,
                "NA",
                "NA",
                "NA",
            ]  # file_name / PASS or FAIL / HW Training / PMAD / PCS
        elif TestItem == "proteanTecs_Test":
            if self.eye1D2D == "NA":
                print(f"Training 1D Eye Diagram")
                self.eye_scan = "1d"
                self.eye_graph_en = 0
                self.vref_start = "0x00"
                self.HW_Training_init()
                self.ChkLog_fail(find="failed")
                self.Test_Info_list = [
                    str(self.pass_fail),
                    self.pll_LOL,
                    self.PASS_FAIL_HW_chk(),
                    "NA",
                    "NA",
                ]  # file_name / PASS or FAIL / HW Training / PMAD / PCS
        else:
            pass

        if self.pass_fail == "FAIL":
            color = wx.RED
        else:
            color = (0, 150, 0, 255)
        self.TestItem_Now2_wx.SetBackgroundColour(color)
        self.TestItem_Now2_wx.Value = (
            f"{self.TestItem_full} , \nTest Result : {self.pass_fail}"
        )

    def PLL_Checking_init(self):
        self.pll_LOL = self.run_0.PLL_Checking(
            pll_en_reg=self.pll_en_reg,
            mode=self.Chip_Mode,
            TestItem=self.TestItem,
            data_rate=float(self.TestDataRate) * 1000,
        )
        # tx_die = Glink_run.g_tx_die
        # rx_die = Glink_run.g_rx_die

    def HW_Training_init(self):
        data_training_en = 1
        self.sys_rst_num = 2
        self.GUC_chip_rst()
        self.PLL_Checking_init()
        print("Start Test M4_D0V_D1V_mode")
        self.slice_result = self.run_0.Hardware_Training_Non(
            mode="M4_D0V_D1V_mode",
            TestItem=self.TestItem,
            data_rate=float(self.TestDataRate) * 1000,
            data_training_en=data_training_en,
            hw_non_1=self.hw_non_1,
            setup_lane=self.even_6,
            vref_start=self.vref_start,
            log_type=self.log_type,
            eye_scan=self.eye_scan,
        )
        self.data_replay = 0
        self.bist = self.run_0.PCS_BIST_Check_NON(
            mode="M4_D0V_D1V_mode",
            PCS_BIST_Check_NON=self.PCS_BIST_Check_NON,
            chk_loop=self.BIST_loop,
            chk_time=self.BIST_time,
            data_replay=self.data_replay,
            voltage_sense_avdd=self.voltage_sense_avdd,
            avdd_sense_en=self.avdd_sense_en,
        )
        self.run_0.proteantecs(mode=0)
        # self.run_0.proteantecs()

        print("\n\nStart Test M4_D1H_D2V_mode")
        self.GUC_chip_rst()
        self.PLL_Checking_init()
        self.slice_result = self.run_0.Hardware_Training_Non(
            mode="M4_D1H_D2V_mode",
            TestItem=self.TestItem,
            data_rate=float(self.TestDataRate) * 1000,
            data_training_en=data_training_en,
            hw_non_1=self.hw_non_1,
            setup_lane=self.even_6,
            vref_start=self.vref_start,
            lane_set_arr=self.lane_set_arr,
            log_type=self.log_type,
            eye_scan=self.eye_scan,
        )
        self.data_replay = 0
        self.bist = self.run_0.PCS_BIST_Check_NON(
            mode="M4_D1H_D2V_mode",
            PCS_BIST_Check_NON=self.PCS_BIST_Check_NON,
            chk_loop=self.BIST_loop,
            chk_time=self.BIST_time,
            data_replay=self.data_replay,
            voltage_sense_avdd=self.voltage_sense_avdd,
            avdd_sense_en=self.avdd_sense_en,
        )
        self.run_0.proteantecs(mode=1)
        # self.run_0.proteantecs()

        self.eye_scan_en = 0

    def SW_Training_init(self):
        bist_mode_select = "pmad"
        if self.sw_vref_type == "All_H":
            all_result = self.run_0.Software_Training_Non_H(
                mode=self.Chip_Mode,
                PCS_BIST_Check_NON=self.PCS_BIST_Check_NON,
                PMAD_BIST_Check_NON=self.PMAD_BIST_Check_NON,
                bist_mode_select=bist_mode_select,
                set_vref=self.set_vref_now,
                pi_step_even=self.pi_step_even,
                vref_loop_now=self.vref_loop_now,
                sw_vref_type=self.sw_vref_type,
                chk_time=int(self.BIST_time),
                sw_vref_loops=self.sw_vref_loops,
                vref_en=self.vref_en,
                lane_valid_en=self.lane_valid_en,
                vref_bypass=self.vref_bypass,
                sw_bist=self.sw_bist,
            )
        else:
            all_result = self.run_0.Software_Training_Non_W(
                mode=self.Chip_Mode,
                PCS_BIST_Check_NON=self.PCS_BIST_Check_NON,
                PMAD_BIST_Check_NON=self.PMAD_BIST_Check_NON,
                bist_mode_select=bist_mode_select,
                set_vref=self.set_vref_now,
                pi_step_even=self.pi_step_even,
                vref_loop_now=self.vref_loop_now,
                sw_vref_type=self.sw_vref_type,
                chk_time=int(self.BIST_time),
                sw_vref_loops=self.sw_vref_loops,
                vref_en=self.vref_en,
                lane_valid_en=self.lane_valid_en,
                vref_bypass=self.vref_bypass,
                sw_bist=self.sw_bist,
                voltage_sense_avdd=self.voltage_sense_avdd,
                avdd_sense_en=self.avdd_sense_en,
            )
        # self.Save_i2cLog(log_name='[Sequence] Software_Training Done\n')
        return all_result

    def PCS_BIST_Check_NON_Path(self):
        self.bist = self.run_0.PCS_BIST_Check_NON(
            mode=self.Chip_Mode,
            PCS_BIST_Check_NON=self.PCS_BIST_Check_NON,
            chk_loop=self.BIST_loop,
            chk_time=self.BIST_time,
            data_replay=self.data_replay,
            voltage_sense_avdd=self.voltage_sense_avdd,
            avdd_sense_en=self.avdd_sense_en,
        )

    def PMAD_BIST_Check_NON_Path(self):
        self.bist = self.run_0.PMAD_BIST_Check_NON(
            mode=self.Chip_Mode,
            PMAD_BIST_Check_NON=self.PMAD_BIST_Check_NON,
            chk_time=self.BIST_time,
            chk_loop=self.BIST_loop,
            voltage_sense_avdd=self.voltage_sense_avdd,
            avdd_sense_en=self.avdd_sense_en,
        )

    def PMAD_BIST_Check_NON_Path_Loop1(self):
        self.bist = self.run_0.PMAD_BIST_Check_NON(
            mode=self.Chip_Mode,
            PMAD_BIST_Check_NON=self.PMAD_BIST_Check_NON,
            chk_time=1,
            chk_loop=1,
            voltage_sense_avdd=self.voltage_sense_avdd,
            avdd_sense_en=self.avdd_sense_en,
        )

    def PMAD_BIST_Check_NLB_Path(self):
        pass

    def Start_Test(self, **kwargs):
        run_n = kwargs.get("run_n", 0)

        self.slice_result = [""]
        S_time = datetime.datetime.now()
        buffer = (os.getcwd()).split("\\")
        tools_version = buffer[len(buffer) - 1]
        print(f"AutoTest Tools Version : {tools_version}")

        print("\n< Start Test >", flush=True)
        color = (0, 0, 0, 0)
        reg_save = self.reg_sequence.Value
        self.run_0.log_info(
            f"Test Tempature={self.Temp_now} Degree_DataRate={self.TestDataRate} Gb/s_Test Mode={self.Chip_Mode}"
        )
        self.phy_0.log_info(
            f"Test Tempature={self.Temp_now} Degree_DataRate={self.TestDataRate} Gb/s_Test Mode={self.Chip_Mode}",
            reg_save,
        )
        # clear log
        textfile = open("TestTools/Graph_Result.txt", "w")
        textfile.write("")
        textfile.close()
        textfile = open("TestTools/Graph_Eye.txt", "w")
        textfile.write("")
        textfile.close()
        textfile = open("TestTools/i2c_log.txt", "w")
        textfile.write("")
        textfile.close()

        self.TestItem_Now2_wx.SetBackgroundColour(color)
        self.TestItem_Now_wx.Value = f"{self.Chip_Mode}_{self.TestDataRate}Gbps"
        self.TestItem_Now2_wx.Value = self.TestItem
        self.Step_count.Range = self.auto_on_num * int(self.Test_Cycle_wx.Value)
        self.Step_count.Value = self.Test_step_Now
        self.m_textCtrl9.Value = f"Test Item : {self.Test_step_Now} of {self.auto_on_num * int(self.Test_Cycle_wx.Value)}  /  Test Cycle : {run_n + 1} of {self.Test_Cycle_wx.Value}"

        if self.abp_pass_fail == "PASS":
            if self.Register_init_en == 1:
                self.Register_init()
            self.TestItem_Init(TestItem=self.TestItem)
        else:
            self.TestItem_Init(TestItem="abp_failed")
            self.TestResult = ["abp_failed"]

        # save test log.txt
        f = open("TestTools/i2c_log.txt", "r")
        i2c_log = f.read()
        f.close()

        print("Elapsed 1 Item : ", datetime.datetime.now() - S_time, flush=True)
        print("\n")
        self.write_log(self.m_richText1.Value + i2c_log)  # save test Sequence
        self.txt_line = int(self.total_lines(self.save_log)) - 5
        self.i2c_txt_line = int(self.total_lines("TestTools/i2c_log.txt"))

        # if self.bypass_report != 1:
        #     self.xlsx_Report()
        # # if self.sw_en == 1:
        # #     self.xlsx_log_link()
        #
        # # self.info_window_wx.Selection=0
        # self.eye_graph_info_wx.SetBitmap(wx.Bitmap(u"TestTools/TestTools_GUI.png", wx.BITMAP_TYPE_ANY))
        self.m_textCtrl9.Value = "Test Done"

        # # for tsmc demo use
        # path = 'C:/Users\GUC\Desktop\D2D AutoTest V01.01_Beta20250328\\\Test Report\\\Test Report Log\Demo.txt'
        # textfile = open(path , "a+")
        # textfile.write(self.m_richText1.Value + i2c_log )
        # textfile.close()
        #
        #
        # textfile = open("TestTools/Graph_Eye1.txt", "w")
        # textfile.write('')
        #
        # f = open(
        #     'Test Report\\\Test Report Log/Demo.txt','r')
        # log = (f.read())
        # log_arr = log.split('\n')
        #
        # textfile = open("TestTools/Graph_Eye1.txt", "a+")
        #
        # for i in range(64):
        #     buffer = (log_arr[i + 10776])[30:75]
        #     textfile.write(buffer + '\n')
        # textfile.close()
        # # print('Done')
        #
        # os.system("Graph.exe")
        # time.sleep(5)
        # command = "taskkill /f /t /im Graph.exe"
        # os.system(command)

    def Start_Test_Specialized(self, **kwargs):
        run_n = kwargs.get("run_n", 0)

        S_time = datetime.datetime.now()
        buffer = (os.getcwd()).split("\\")
        tools_version = buffer[len(buffer) - 1]
        print(f"AutoTest Tools Version : {tools_version}")

        print("\n< Start Specialized Test >", flush=True)
        color = (0, 0, 0, 0)
        self.TestItem_Now2_wx.SetBackgroundColour(color)
        self.TestItem_Now_wx.Value = f"{self.TestDataRate}Gbps"
        self.TestItem_Now2_wx.Value = self.TestItem_full
        self.Step_count.Range = self.auto_on_num * int(self.Test_Cycle_wx.Value)
        self.Step_count.Value = self.Test_step_Now
        self.m_textCtrl9.Value = f"Test Item : {self.Test_step_Now} of {self.auto_on_num * int(self.Test_Cycle_wx.Value)}  /  Test Cycle : {run_n + 1} of {self.Test_Cycle_wx.Value}"
        self.TestItem_Init(TestItem=self.TestItem)

        print("Elapsed 1 Item : ", datetime.datetime.now() - S_time, flush=True)
        self.write_log(self.m_richText1.Value)  # save test Sequence

        self.info_window_wx.Selection = 1
        self.eye_graph_info_wx.SetBitmap(
            wx.Bitmap("TestTools/TestTools_GUI.png", wx.BITMAP_TYPE_ANY)
        )

    def log_name(self):
        Loop_time = datetime.datetime.now()
        Loop_time = ((str(Loop_time)).replace(":", "-"))[0:19]

        if self.TestItem == "VCO":
            self.Log_Folder_path = (
                self.TestItem_full
                + "_"
                + str(self.Temp_now)
                + "Degree_"
                + self.chip_version
                + "_"
                + str(self.TestDataRate)
                + "Gbps_"
                + str(self.Chip_Mode)
                + "_"
                + str(self.eye_scan)
                + "_"
                + str(self.File_name)
                + "_"
                + str(Loop_time)
            )
        else:
            if self.File_name == "NA" or self.File_name == None:
                self.Log_Folder_path = (
                    self.TestItem_full
                    + "_"
                    + str(self.Temp_now)
                    + "Degree_"
                    + self.chip_version
                    + "_"
                    + str(self.TestDataRate)
                    + "Gbps_"
                    + str(self.Chip_Mode)
                    + "_"
                    + str(self.eye_scan)
                    + "_"
                    + str(self.File_name)
                    + "_"
                    + str(Loop_time)
                )
            else:
                self.Log_Folder_path = (
                    self.TestItem_full
                    + "_"
                    + str(self.Temp_now)
                    + "Degree_"
                    + self.chip_version
                    + "_"
                    + str(self.TestDataRate)
                    + "Gbps_"
                    + str(self.Chip_Mode)
                    + "_"
                    + str(self.eye_scan)
                    + "_"
                    + str(self.File_name)
                    + "_"
                    + str(Loop_time)
                )

        self.save_log = "Test Report/Test Report Log/" + self.Log_Folder_path + ".txt"
        self.log_path = "Test Report Log/" + self.Log_Folder_path + ".txt"
        self.graph_info = f"{self.TestItem_full} {self.chip_version} {self.Temp_now}Degree C {self.TestDataRate}Gb/s {self.Chip_Mode}"

        # save all slice eye width
        if self.eye1D2D != "NA":
            for n in range(12):
                slice_path = f"TestTools/{self.eye_txt_arr[n]}"
                f = open(slice_path, "r")
                i2c_log = f.read()
                f.close()
                path = f"TestTools/D0_D1_D2_2D_Slice.txt"
                textfile = open(path, "a+")
                textfile.write(self.eye_txt_arr[n])
                textfile.close()
                textfile = open(path, "a+")
                textfile.write(i2c_log)
                textfile.close()
                print("\n\n")
                print(self.eye_txt_arr[n])
                print(i2c_log)

                textfile = open(slice_path, "a+")
                textfile.write(
                    f"{self.Chip_Mode}\n{self.TestDataRate}\n{self.eye_txt_arr[n]}\n{self.save_log}"
                )
                textfile.close()

    def GUC_chip_rst(self):
        if self.sys_rst_num == 0:  # GPIO Reset
            print("\nGPIO Reset Test Chip Reset\n", flush=True)
            self.phy_0.pico_gpio_low(6, 1)
            self.phy_0.pico_gpio_low(6, 0)
            self.phy_0.pico_gpio_low(6, 1)
            self.phy_0.pico_gpio_low(7, 1)
            self.phy_0.pico_gpio_low(7, 0)
            self.phy_0.pico_gpio_low(7, 1)
            self.phy_0.pico_gpio_low(8, 1)
            self.phy_0.pico_gpio_low(8, 0)
            self.phy_0.pico_gpio_low(8, 1)
        elif self.sys_rst_num == 1:  # Power Reset
            print("\nPower Cycle Test Chip Reset\n", flush=True)
            self.visa.E36233A_Out_OFF_RST()
            time.sleep(2)
            self.visa.E36233A_Out_ON_RST()
            time.sleep(2)
        elif self.sys_rst_num == 2:  # PMIC Power Reset
            print("\nPASS\n", flush=True)
            # print('\nPower PMIC Reset\n', flush=True)
            # self.phy_0.TPSM831D31_Output_Disable(0x1, 'CHA')  # VDD
            # self.phy_0.TPSM831D31_Output_Disable(0x1, 'CHB')  # IOVDD
            # self.phy_0.TPSM831D31_Output_Disable(0x2, 'CHA')  # D0_AVDD_V2
            # self.phy_0.TPSM831D31_Output_Disable(0x2, 'CHB')  # D0_AVDD12_V2
            # self.phy_0.TPSM831D31_Output_Disable(0x4, 'CHA')  # D0_AVDD_V1_D1_V1
            # self.phy_0.TPSM831D31_Output_Disable(0x4, 'CHB')  # D0_AVDD12_V1_D1_V1
            # self.phy_0.TPSM831D31_Output_Disable(0x8, 'CHA')  # D1_AVDD_V2_D2_V1
            # self.phy_0.TPSM831D31_Output_Disable(0x8, 'CHB')  # D1_AVDD12_V2_D2_V1
            # self.phy_0.TPSM831D31_Output_Disable(0x10, 'CHA')  # D2_AVDD_V2
            # self.phy_0.TPSM831D31_Output_Disable(0x10, 'CHB')  # D2_AVDD12_V2
            # time.sleep(1)
            # self.phy_0.TPSM831D31_Output_Enable(0x1, 'CHA')  # VDD
            # self.phy_0.TPSM831D31_Output_Enable(0x1, 'CHB')  # IOVDD
            # self.phy_0.TPSM831D31_Output_Enable(0x2, 'CHA')  # D0_AVDD_V2
            # self.phy_0.TPSM831D31_Output_Enable(0x2, 'CHB')  # D0_AVDD12_V2
            # self.phy_0.TPSM831D31_Output_Enable(0x4, 'CHA')  # D0_AVDD_V1_D1_V1
            # self.phy_0.TPSM831D31_Output_Enable(0x4, 'CHB')  # D0_AVDD12_V1_D1_V1
            # self.phy_0.TPSM831D31_Output_Enable(0x8, 'CHA')  # D1_AVDD_V2_D2_V1
            # self.phy_0.TPSM831D31_Output_Enable(0x8, 'CHB')  # D1_AVDD12_V2_D2_V1
            # self.phy_0.TPSM831D31_Output_Enable(0x10, 'CHA')  # D2_AVDD_V2
            # self.phy_0.TPSM831D31_Output_Enable(0x10, 'CHB')  # D2_AVDD12_V2
        else:
            # Skip Reset
            print("\nSkip Test Chip Reset\n", flush=True)
            pass
        self.phy_0.resetn(abp_en=1)

    def read_pll_write_map(self, **kargs):
        pll_pmaa_map = kargs.get("pll_pmaa_map", "")

        array = []
        for r in range(len(pll_pmaa_map)):
            reg_list = (pll_pmaa_map[r]).split(",")
            offset = (((reg_list[2]).split("_"))[1]).lstrip()
            bit = reg_list[3]

            if reg_list[5].find("Die") != -1:
                reg_value = f"0x{offset},{bit},0x{reg_list[5]}"
            elif reg_list[5].find("nan") != -1:
                reg_value = f"0x{offset},{bit},0x{reg_list[5]}"
            elif reg_list[5].find("-") != -1:
                reg_value = f"0x{offset},{bit},0x{reg_list[5]}"
            elif reg_list[5].find("SN") != -1:
                reg_value = f"0x{offset},{bit},0x{reg_list[5]}"
            elif reg_list[5].find("EW") != -1:
                reg_value = f"0x{offset},{bit},0x{reg_list[5]}"
            elif reg_list[5].find("h") != -1:
                register_value = (reg_list[5]).split("h")
                register_value = register_value[1]
                reg_value = f"0x{offset},{bit},0x{register_value}"
            else:
                reg_value = f"0x{offset},{bit},0x{reg_list[5]}"
                print(f"\033PLL Register Map Read Failed{reg_value}", flush=True)
            # print(reg_value)
            array += [reg_value]
        return array

    """Register Read/Write"""

    def reg_map_load_event(self, event):
        self.phy_0.indirect_enable(0, 1)
        self.phy_0.indirect_enable(0, 2)
        self.phy_0.indirect_enable(1, 1)
        self.phy_0.indirect_enable(1, 2)
        self.phy_0.indirect_enable(2, 1)
        self.phy_0.indirect_enable(2, 2)

        self.reg_map_load.Value = True
        self.reg_map_load.Label = "Running"
        seach_str = self.tree_seach.GetValue()

        self.tree_item.DeleteAllItems()
        buffer = self.pll_map_data()
        buffer1 = self.slice_map_data()
        if buffer == "Ready":
            self.root = self.tree_item.GetRootItem()
            self.r2 = self.tree_item.AppendItem(self.root, "PLL_Register")
            for i in range(len(self.pll_map_tree)):
                reg_list = (self.pll_map_tree)[i].split(",")
                if reg_list[0].find(seach_str) != -1:
                    child = self.tree_item.AppendItem(self.r2, "1")
                    self.tree_item.SetItemText(child, 0, f"{reg_list[0]}")  # Name
                    self.tree_item.SetItemText(child, 1, f"{reg_list[1]}")  # offset
                    self.tree_item.SetItemText(child, 2, f"{reg_list[2]}")  # bit
                    self.tree_item.SetItemText(child, 3, f"{reg_list[3]}")  # map value
                    self.tree_item.SetItemText(child, 4, f"PLL")  # Map type
                    self.tree_item.SetItemText(child, 5, f"{reg_list[4]}")  # Note
                    self.tree_item.Expand(child)
                else:
                    pass

        if buffer1 == "Ready":
            self.root = self.tree_item.GetRootItem()
            self.r2 = self.tree_item.AppendItem(self.root, "Slice_Register")
            for i in range(len(self.slice_map_tree)):
                reg_list = (self.slice_map_tree)[i].split(",")
                if reg_list[0].find(seach_str) != -1:
                    child = self.tree_item.AppendItem(self.r2, "1")
                    self.tree_item.SetItemText(child, 0, f"{reg_list[0]}")  # Name
                    self.tree_item.SetItemText(child, 1, f"{reg_list[1]}")  # offset
                    self.tree_item.SetItemText(child, 2, f"{reg_list[2]}")  # bit
                    self.tree_item.SetItemText(child, 3, f"{reg_list[3]}")  # map value
                    self.tree_item.SetItemText(child, 4, f"{reg_list[5]}")  # Map type
                    self.tree_item.SetItemText(child, 5, f"{reg_list[4]}")  # Note
                    self.tree_item.Expand(child)
                else:
                    pass

        self.reg_map_load.Value = False
        self.reg_map_load.Label = "Press Load Register Map"

    def reg_compare_event(self, event):
        self.m_richText1.Clear()
        self.phy_0.indirect_enable(0, 1)
        self.phy_0.indirect_enable(0, 2)
        self.phy_0.indirect_enable(1, 1)
        self.phy_0.indirect_enable(1, 2)
        self.phy_0.indirect_enable(2, 1)
        self.phy_0.indirect_enable(2, 2)

        self.reg_map_load.Value = True
        self.reg_compare.Label = "Compare Now"
        seach_str = self.tree_seach.GetValue()

        buffer = self.pll_map_data()
        buffer1 = self.slice_map_data()
        die_sel = self.die_select.GetValue()
        group_sel = self.group_select.GetValue()
        slice_sel = self.slice_select.GetValue()
        pll_compare_info = f"Die{die_sel}{group_sel}"

        if buffer == "Ready":
            pll_compare_data = []
            for i in range(len(self.pll_map_tree)):
                reg_list = (self.pll_map_tree)[i].split(",")
                if reg_list[0].find(seach_str) != -1:
                    pll_name = f"{reg_list[0]}"
                    self.offset = f"{reg_list[1]}"
                    self.bit = f"{reg_list[2]}"
                    if self.bit.find(":") != -1:
                        bit_list = self.bit.split(":")
                        self.i2c_bit_MSB.Value = bit_list[0]
                        self.i2c_bit_LSB.Value = bit_list[1]
                    else:
                        self.i2c_bit_MSB.Value = self.bit
                        self.i2c_bit_LSB.Value = self.bit
                    pll_map_value = f"{reg_list[3]}"
                    pll_note = f"{reg_list[4]}"

                    self.reg_source.Value = "User_Define"
                    self.i2c_read_compare()
                    if int(pll_map_value, 16) == int(self.compare_value, 16):
                        pll_compare_value = "PASS"
                    else:
                        pll_compare_value = "Failed"
                    # pll_compare_result = (f'{pll_name} = {pll_offset}, Bit=[{pll_bit}], Map_Value={pll_map_value}, Read Value={self.i2c_register_value.Value}, Compare_Value={pll_compare_value}')
                    # print(pll_compare_result)
                    pll_compare_result = f"{self.offset},Bit=[{self.bit}],{pll_name},{pll_note},PLL,{pll_map_value},{self.i2c_register_value.Value},{pll_compare_value},{pll_compare_info}"
                    pll_compare_data.append(pll_compare_result)
                else:
                    pass

        slice_compare_info = f"Die{die_sel}{group_sel}_Slice{slice_sel}"
        slice_compare_data = []
        if buffer1 == "Ready":
            for i in range(len(self.slice_map_tree)):
                reg_list = (self.slice_map_tree)[i].split(",")
                if reg_list[0].find(seach_str) != -1:
                    slice_name = f"{reg_list[0]}"
                    self.offset = f"{reg_list[1]}"
                    self.bit = f"{reg_list[2]}"
                    slice_map_value = f"{reg_list[3]}"
                    slice_note = f"{reg_list[4]}"

                    self.reg_source.Value = "User_Define"
                    self.i2c_read_compare()
                    if int(slice_map_value, 16) == int(self.compare_value, 16):
                        slice_compare_value = "PASS"
                    else:
                        slice_compare_value = "Failed"

                    slice_compare_result = f"{self.offset},Bit=[{self.bit}],{slice_name},{slice_note},Slice,{slice_map_value},{self.i2c_register_value.Value},{slice_compare_value},{slice_compare_info}"
                    slice_compare_data.append(slice_compare_result)
                else:
                    pass

        buffer2 = pll_compare_data + slice_compare_data
        self.xlsx_reg_compare(buffer2)
        self.reg_map_load.Value = False
        self.reg_compare.Label = "Run Compare"

    def tree_seach_event(self, event):
        seach_str = self.tree_seach.GetValue()

        self.tree_item.DeleteAllItems()
        buffer = self.pll_map_data()
        buffer1 = self.slice_map_data()
        if buffer == "Ready":
            self.root = self.tree_item.GetRootItem()
            self.r2 = self.tree_item.AppendItem(self.root, "PLL_Register")
            for i in range(len(self.pll_map_tree)):
                reg_list = (self.pll_map_tree)[i].split(",")
                if reg_list[0].find(seach_str) != -1:
                    child = self.tree_item.AppendItem(self.r2, "1")
                    self.tree_item.SetItemText(child, 0, f"{reg_list[0]}")
                    self.tree_item.SetItemText(child, 1, f"{reg_list[1]}")
                    self.tree_item.SetItemText(child, 2, f"{reg_list[2]}")
                    self.tree_item.SetItemText(child, 3, f"{reg_list[3]}")
                    self.tree_item.SetItemText(child, 4, f"{reg_list[4]}")
                    self.tree_item.Expand(child)
                else:
                    pass

        if buffer1 == "Ready":
            self.root = self.tree_item.GetRootItem()
            self.r2 = self.tree_item.AppendItem(self.root, "Slice_Register")
            for i in range(len(self.slice_map_tree)):
                reg_list = (self.slice_map_tree)[i].split(",")
                if reg_list[0].find(seach_str) != -1:
                    child = self.tree_item.AppendItem(self.r2, "1")
                    self.tree_item.SetItemText(child, 0, f"{reg_list[0]}")
                    self.tree_item.SetItemText(child, 1, f"{reg_list[1]}")
                    self.tree_item.SetItemText(child, 2, f"{reg_list[2]}")
                    self.tree_item.SetItemText(child, 3, f"{reg_list[3]}")
                    self.tree_item.SetItemText(child, 4, f"{reg_list[4]}")
                    self.tree_item.SetItemText(child, 5, f"{reg_list[5]}")
                    self.tree_item.Expand(child)
                else:
                    pass

    def select_tree_item_event(self, event):
        the_one = self.tree_item.GetSelection()
        parent_now = self.tree_item.GetItemParent(the_one)
        parent_name = self.tree_item.GetItemText(parent_now)
        # print(parent_name)

        reg_name = self.tree_item.GetItemText(the_one, col=0)
        offset = self.tree_item.GetItemText(the_one, col=1)
        bit = self.tree_item.GetItemText(the_one, col=2)
        value = self.tree_item.GetItemText(the_one, col=3)
        # print(f'{reg_name}, {offset}, {bit}, {value}, {note}')

        # Read register value
        if bit.find(":") == -1:
            self.i2c_bit_MSB.Value = str(bit)
            self.i2c_bit_LSB.Value = str(bit)
        else:
            buffer = bit.split(":")
            self.i2c_bit_MSB.Value = buffer[0]
            self.i2c_bit_LSB.Value = buffer[1]
        self.offset_ref = "tree"
        self.offset = self.i2c_offset.Value = offset
        self.bit = bit
        self.i2c_read(self)

    def pll_map_data(self):
        if self.pll_sheet_name_tree.GetValue() != "":
            self.map_col_name = self.pll_sheet_col_num.GetValue()
            self.map_print_en = 0
            pll_pmaa_map = self.read_xls_tree(map="pll", map_path=self.pll_map_path)
            self.pll_map_tree = self.read_pll_write_map_tree(pll_pmaa_map=pll_pmaa_map)
            buffer = "Ready"
        else:
            # h = ctypes.windll.LoadLibrary("C:\\Windows\\System32\\user32.dll")
            # h.MessageBoxW(0, u'Please Select  " PLL Map Sheet Select "', u'Warning !! ', 0)
            buffer = "Warning"
        return buffer

    def slice_map_data(self):
        if self.slice_sheet_name_tree.GetValue() != "":
            self.map_col_name = self.slice_sheet_col_num.GetValue()
            self.map_print_en = 0
            slice_pmaa_map = self.read_xls_tree(
                map="slice", map_path=self.slice_map_path
            )
            self.slice_map_tree = self.read_slice_write_map_tree(
                slice_pmaa_map=slice_pmaa_map
            )
            buffer = "Ready"
        else:
            # h = ctypes.windll.LoadLibrary("C:\\Windows\\System32\\user32.dll")
            # h.MessageBoxW(0, u'Please Select  " Slice Map Sheet Select "', u'Warning !! ', 0)
            buffer = "Warning"
        return buffer

    def read_xls_tree(self, **kwargs):
        map = kwargs.get("map", "")

        col_arr = [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "X",
            "Y",
            "Z",
        ]
        for col in range(26):
            if col_arr[col] == self.map_col_name:
                sheet_col = col
            else:
                pass
        if map == "pll":
            xls = pd.ExcelFile(self.pll_map_path)
            # print(xls.sheet_names)

            sheet_num = self.pll_sheet_name_tree.GetSelection()
            df = xls.parse(sheet_num)
            col_num = len(df.axes[0])

            reg_arr = []
            for r1 in range(col_num):
                buffer1 = df.loc[r1]
                group = "tport"
                name = str(buffer1[2])  # register name

                offset = str(buffer1[0])  # register offset
                if offset == " ":
                    for q in range(99):
                        buffer2 = df.loc[r1 - q]
                        offset = str(buffer2[0])  # register offset
                        if offset != " ":
                            break
                        else:
                            pass
                else:
                    offset = str(buffer1[0])  # register offset
                offset = f"0x{offset}"

                bit = str(buffer1[1])  # register bit
                bit = bit.replace("[", "")
                bit = bit.replace("]", "")

                RW = str(buffer1[3])  # register RW

                reg_value = str(buffer1[sheet_col])  # register reg_value
                # if reg_value.find('h') !=-1:
                #     reg_value = (reg_value.split(('h')))[1]
                #     reg_value = f'0x{reg_value}'
                #     reg_value = str(int(reg_value, 16))
                # elif reg_value.find('d') !=-1:
                #     reg_value = (reg_value.split(('d')))[1]
                # else:
                #     reg_value = 'error'
                # reg_value = str(buffer1[5])  # register reg_value(chip default value)
                reg = f"{group}, {name}, {offset}, {bit}, {RW}, {reg_value}"
                reg_arr += [reg]
        else:
            pass

            # Tony pll map format use
            # reg_arr = []
            # for r1 in range(col_num-1):
            #     buffer1 = df.iloc[r1]
            #     group = str(buffer1[0])  # group
            #     name = str(buffer1[1])  # register name
            #     if group == 'PLL' or group == 'SLICE0':
            #         offset = str(buffer1[2])  # register offset
            #         bit = str(buffer1[3])  # register bit
            #         RW = str(buffer1[4])  # register RW
            #         reg_value = str(buffer1[sheet_col])  # register reg_value
            #         # reg_value = str(buffer1[5])  # register reg_value(chip default value)
            #         reg = (f'{group}, {name}, {offset}, {bit}, {RW}, {reg_value}')
            #         reg_arr += [reg]
            #     else:
            #         pass
        if map == "slice":
            xls = pd.ExcelFile(self.slice_map_path)
            # print(xls.sheet_names)

            sheet_num = self.slice_sheet_name_tree.GetSelection()
            df = xls.parse(sheet_num)
            col_num = len(df.axes[0])

            reg_arr = []
            for r1 in range(col_num):
                buffer1 = df.loc[r1]
                group = "tport"
                name = str(buffer1[2])  # register name

                offset = str(buffer1[0])  # register offset
                if offset == " ":
                    for q in range(99):
                        buffer2 = df.loc[r1 - q]
                        offset = str(buffer2[0])  # register offset
                        if offset != " ":
                            break
                        else:
                            pass
                else:
                    offset = str(buffer1[0])  # register offset
                offset = f"0x{offset}"

                bit = str(buffer1[1])  # register bit
                bit = bit.replace("[", "")
                bit = bit.replace("]", "")

                RW = str(buffer1[3])  # register RW

                reg_value = str(buffer1[sheet_col])  # register reg_value

                type = str(buffer1[6])  # register RW
                # if reg_value.find('h') !=-1:
                #     reg_value = (reg_value.split(('h')))[1]
                #     reg_value = f'0x{reg_value}'
                #     reg_value = str(int(reg_value, 16))
                # elif reg_value.find('d') !=-1:
                #     reg_value = (reg_value.split(('d')))[1]
                # else:
                #     reg_value = 'error'
                # reg_value = str(buffer1[5])  # register reg_value(chip default value)
                reg = f"{group}, {name}, {offset}, {bit}, {RW}, {reg_value}, {type}"
                reg_arr += [reg]
        else:
            pass
        return reg_arr

    def read_pll_write_map_tree(self, **kargs):
        pll_pmaa_map = kargs.get("pll_pmaa_map", "")

        array = []
        for r in range(len(pll_pmaa_map)):
            reg_list = (pll_pmaa_map[r]).split(",")
            offset = hex(
                int(reg_list[2], 16) + int(0x2000)
            )  # pll regisyer start 0x2000
            bit = reg_list[3]
            if reg_list[5].find("h") != -1:
                register_value = (reg_list[5]).split("h")
                register_value = int(register_value[1], 16)
            elif reg_list[5].find("d") != -1:
                register_value = (reg_list[5]).split("d")
                register_value = register_value[1]
            elif reg_list[5].find("b") != -1:
                register_value = (reg_list[5]).split("b")
                register_value = register_value[1]
            else:
                register_value = "register_value_error"
            register_value = hex(int(register_value))

            reg_value = f"{reg_list[1]},{offset},{bit},{register_value},{reg_list[4]}"
            array += [reg_value]
        return array

        # Tony pll map format use
        # array=[]
        # for r in range(len(pll_pmaa_map)):
        #     reg_list = (pll_pmaa_map[r]).split(",")
        #     offset = (((reg_list[2]).split('_'))[1]).lstrip()
        #     bit = reg_list[3]
        #
        #     if reg_list[5].find('Die') != -1:
        #         pass
        #     elif reg_list[5].find('nan') != -1:
        #         reg_value =  f'{reg_list[1]},{offset},{bit},{reg_list[5]},{reg_list[4]}'
        #     elif reg_list[5].find('-') != -1:
        #
        #         reg_value = f'{reg_list[1]},0x{offset},{bit},0x{reg_list[5]},{reg_list[4]}'
        #     elif reg_list[5].find('SN') != -1:
        #         aaa='aaa'
        #         reg_value = f'{reg_list[1]},0x{offset},{bit},0x{aaa},{reg_list[4]}'
        #     elif reg_list[5].find('EW') != -1:
        #         aaa='aaa'
        #         reg_value = f'{reg_list[1]},0x{offset},{bit},0x{aaa},{reg_list[4]}'
        #     elif reg_list[5].find('h') != -1:
        #         register_value = (reg_list[5]).split("h")
        #         register_value = register_value[1]
        #         reg_value = f'{reg_list[1]},0x{offset},{bit},0x{register_value},{reg_list[4]}'
        #     else:
        #         pass
        #     array += [reg_value]
        # return array

    def read_slice_write_map_tree(self, **kargs):
        slice_pmaa_map = kargs.get("slice_pmaa_map", "")

        array = []
        for r in range(len(slice_pmaa_map)):
            reg_list = (slice_pmaa_map[r]).split(",")
            type = re.sub(r"\s+", "", reg_list[6])
            if type == "UCIe":
                buffer = 0
            elif type == "Slice":
                buffer = int(0x3000)
            elif type == "PCS":
                buffer = int(0x7000)
            elif type == "Adapter":
                buffer = int(0x7800)
            else:
                print("Register offset error , failed")
            offset = hex(
                int(reg_list[2], 16) + buffer
            )  # UCIe/Slice/PCS/Adapter regisyer start
            bit = reg_list[3]
            if reg_list[5].find("h") != -1:
                register_value = (reg_list[5]).split("h")
                register_value = int(register_value[1], 16)
            elif reg_list[5].find("d") != -1:
                register_value = (reg_list[5]).split("d")
                register_value = register_value[1]
            elif reg_list[5].find("b") != -1:
                register_value = (reg_list[5]).split("b")
                register_value = register_value[1]
            else:
                register_value = "register_value_error"
            register_value = hex(int(register_value))

            reg_value = (
                f"{reg_list[1]},{offset},{bit},{register_value},{reg_list[4]},{type}"
            )
            array += [reg_value]
        return array

    def i2c_read_compare(self):
        self.offset_ref = "user"
        self.die = int(self.die_select.GetValue())
        self.group_sel_num = self.group_select.GetValue()
        self.slice = self.slice_select.GetValue()
        self.phy_0.die_sel(die=self.die)

        reg_arr = [
            f"{self.offset},{self.bit},nan,nan,nan,nan,v,{self.die},{self.group_sel_num},{self.slice}"
        ]
        if self.reg_type.GetValue() == "Normal I2C":
            buffer = self.phy_0.normal_i2c(
                self.die,
                self.group_sel_num,
                self.offset_hex,
                self.bit_S,
                self.bit_Leng,
                "read",
            )
        elif self.reg_type.GetValue() == "TPORT":
            buffer = self.phy_0.indirect_read(
                self.EHOST[self.die][0], self.offset_hex, self.bit_S, top=1
            )
        else:
            buffer = self.phy_0.reg_user_set(
                reg_arr=reg_arr,
                show=0,
                mode="gui_tree",
                gui_die_sel=self.die,
                gui_group_num=self.group_sel_num,
                gui_slice_num=self.slice,
            )  # PLL/Slice
        self.i2c_register_value.Value = self.offset_tree = self.compare_value = buffer

    def i2c_read(self, event):
        self.offset_ref = "user"
        if (
            len(self.i2c_bit_MSB.GetValue()) != 0
            and len(self.i2c_bit_LSB.GetValue()) != 0
            and len(self.i2c_offset.GetValue()) != 0
        ):
            self.i2c_read_write()
            self.phy_0.die_sel(die=self.die)
            reg_arr = [
                f"{self.offset},{self.bit},nan,nan,nan,nan,v,{self.die},{self.group_sel_num},{self.slice}"
            ]

            if self.reg_type.GetValue() == "Normal I2C":
                buffer = self.phy_0.normal_i2c(
                    self.die,
                    self.group_sel_num,
                    self.offset_hex,
                    self.bit_S,
                    self.bit_Leng,
                    "read",
                )
            elif self.reg_type.GetValue() == "TPORT":
                buffer = self.phy_0.indirect_read(
                    self.EHOST[self.die][0], self.offset_hex, self.bit_S, top=1
                )
            else:
                buffer = self.phy_0.reg_user_set(
                    reg_arr=reg_arr,
                    show=0,
                    mode="gui_tree",
                    gui_die_sel=self.die,
                    gui_group_num=self.group_sel_num,
                    gui_slice_num=self.slice,
                )  # PLL/Slice

            self.i2c_register_value.Value = self.offset_tree = buffer
        else:
            pass
            # print('\033Please check Offset / Bit value or format')

    def i2c_write(self, event):
        if (
            len(self.i2c_bit_MSB.GetValue()) != 0
            and len(self.i2c_bit_LSB.GetValue()) != 0
            and len(self.i2c_offset.GetValue()) != 0
        ):
            self.i2c_read_write()
            self.phy_0.die_sel(die=self.die)
            buffer = self.i2c_register_value.GetValue()
            self.register_value = int(buffer, 16)

            if self.reg_type.GetValue() == "Normal I2C":
                self.phy_0.normal_i2c(
                    self.die,
                    self.group_sel_num,
                    self.offset_hex,
                    self.bit_S,
                    self.bit_Leng,
                    "write",
                    setv=self.register_value,
                )
            elif self.reg_type.GetValue() == "TPORT":
                self.phy_0.indirect_write(
                    self.EHOST[self.die][0],
                    self.offset_hex,
                    self.bit_S,
                    self.register_value,
                    top=1,
                )
            else:
                reg_arr = [
                    f"{self.offset},{self.bit},{hex(self.register_value)},nan,nan,v,nan,{self.die},{self.group_sel_num},{self.slice}"
                ]
                self.phy_0.reg_user_set(
                    reg_arr=reg_arr,
                    show=0,
                    mode="gui_tree",
                    gui_die_sel=self.die,
                    gui_group_num=self.group_sel_num,
                    gui_slice_num=self.slice,
                )  # PLL/Slice)

    def i2c_read_write(self):
        self.die = int(self.die_select.GetValue())
        self.group_sel_num = self.group_select.GetValue()
        self.slice = self.slice_select.GetValue()
        reg_source = self.reg_source.GetValue()

        if reg_source == "Register_Tree":
            the_one = self.tree_item.GetSelection()
            self.map_type = self.tree_item.GetItemText(the_one, col=4)

        if self.abp_en == 1:
            self.phy_0.indirect_enable(0, 1)
            self.phy_0.indirect_enable(0, 2)
            self.phy_0.indirect_enable(1, 1)
            self.phy_0.indirect_enable(1, 2)
            self.phy_0.indirect_enable(2, 1)
            self.phy_0.indirect_enable(2, 2)
            self.abp_en = 1

        if self.offset_ref == "tree":
            pass
        else:
            self.offset = self.i2c_offset.GetValue()
            self.bit = f"{self.i2c_bit_MSB.GetValue()}:{self.i2c_bit_LSB.GetValue()}"
        buffer = self.offset
        self.offset_hex = int(buffer, 16)

        # normal i2c
        if self.i2c_bit_MSB.GetValue() == self.i2c_bit_LSB.GetValue():
            self.bit_S = self.i2c_bit_MSB.GetValue()
            self.bit_Leng = 1
        else:
            self.bit_S = f"{self.i2c_bit_MSB.GetValue()}:{self.i2c_bit_LSB.GetValue()}"
            self.bit_Leng = int(
                (int(self.i2c_bit_MSB.GetValue()) - int(self.i2c_bit_LSB.GetValue()))
                + 1
            )

    def register_compare(self):
        ip_num = self.ip_version_wx.GetSelection()
        print(f"< Start Compare Register : {ip_num} >")
        self.m_richText1.Clear()

    """"Module Board""" ""

    def power_source_autotest(self, **kargs):
        command = kargs.get("command", "PMIC/0.75")
        num = kargs.get("num", 0)

        if command.find("PMIC") != -1:
            volt = float((command.split("/"))[1])
            self.power_mux_value(num, volt)
        elif command.find("36000") != -1:
            self.visa.E36000A_setup_channel(command=command)
        else:
            pass

    def power_update_even(self, even):
        power_arr = [
            self.power_select1.GetValue(),
            self.power_select2.GetValue(),
            self.power_select3.GetValue(),
            self.power_select4.GetValue(),
            self.power_select5.GetValue(),
            self.power_select6.GetValue(),
            self.power_select7.GetValue(),
            self.power_select8.GetValue(),
            self.power_select9.GetValue(),
            self.power_select10.GetValue(),
            self.power_select11.GetValue(),
            self.power_select12.GetValue(),
            self.power_select13.GetValue(),
            self.power_select14.GetValue(),
            self.power_select15.GetValue(),
            self.power_select16.GetValue(),
            self.power_select17.GetValue(),
            self.power_select18.GetValue(),
            self.power_select19.GetValue(),
            self.power_select20.GetValue(),
        ]
        value_arr = [
            self.power_value1.GetValue(),
            self.power_value2.GetValue(),
            self.power_value3.GetValue(),
            self.power_value4.GetValue(),
            self.power_value5.GetValue(),
            self.power_value6.GetValue(),
            self.power_value7.GetValue(),
            self.power_value8.GetValue(),
            self.power_value9.GetValue(),
            self.power_value10.GetValue(),
            self.power_value11.GetValue(),
            self.power_value12.GetValue(),
            self.power_value13.GetValue(),
            self.power_value14.GetValue(),
            self.power_value15.GetValue(),
            self.power_value16.GetValue(),
            self.power_value17.GetValue(),
            self.power_value18.GetValue(),
            self.power_value19.GetValue(),
            self.power_value20.GetValue(),
        ]
        current_arr = [
            self.power_current1.GetValue(),
            self.power_current2.GetValue(),
            self.power_current3.GetValue(),
            self.power_current4.GetValue(),
            self.power_current5.GetValue(),
            self.power_current6.GetValue(),
            self.power_current7.GetValue(),
            self.power_current8.GetValue(),
            self.power_current9.GetValue(),
            self.power_current10.GetValue(),
            self.power_current11.GetValue(),
            self.power_current12.GetValue(),
            self.power_current13.GetValue(),
            self.power_current14.GetValue(),
            self.power_current15.GetValue(),
            self.power_current16.GetValue(),
            self.power_current17.GetValue(),
            self.power_current18.GetValue(),
            self.power_current19.GetValue(),
            self.power_current20.GetValue(),
        ]
        for num in range(10):
            buffer = power_arr[num]
            volt = float(value_arr[num])
            current = current_arr[num]
            if buffer.find("PMIC") != -1:
                self.power_mux_value(num, volt)
            elif buffer.find("Power") != -1:
                buffer_list = buffer.split("_")
                command = (
                    "36000"
                    + "/"
                    + (buffer_list[1])[6]
                    + "/"
                    + (buffer_list[2])[2]
                    + "/"
                    + str(volt)
                    + "/"
                    + str(current)
                )
                self.visa.E36000A_setup_channel(command=command)
            else:
                pass

    def power_mux_value(self, num, volt):
        if num == 0:
            self.phy_0.TPSM831D31_VoltageSet(0x1, "CHA", volt)  # VDD
        if num == 1:
            self.phy_0.TPSM831D31_VoltageSet(0x1, "CHB", volt)  # IOVDD
        if num == 2:
            self.phy_0.TPSM831D31_VoltageSet(0x2, "CHA", volt)  # D0_AVDD_V2
        if num == 3:
            self.phy_0.TPSM831D31_VoltageSet(0x2, "CHB", volt)  # D0_AVDD12_V2
        if num == 4:
            self.phy_0.TPSM831D31_VoltageSet(0x4, "CHA", volt)  # D0_AVDD_V1_D1_V1
        if num == 5:
            self.phy_0.TPSM831D31_VoltageSet(0x4, "CHB", volt)  # D0_AVDD12_V1_D1_V1
        if num == 6:
            self.phy_0.TPSM831D31_VoltageSet(0x8, "CHA", volt)  # D1_AVDD_V2_D2_V1
        if num == 7:
            self.phy_0.TPSM831D31_VoltageSet(0x8, "CHB", volt)  # D1_AVDD12_V2_D2_V1
        if num == 8:
            self.phy_0.TPSM831D31_VoltageSet(0x10, "CHA", volt)  # D2_AVDD_V2
        if num == 9:
            self.phy_0.TPSM831D31_VoltageSet(0x10, "CHB", volt)  # D2_AVDD12_V2
        else:
            pass

    def TPSM831D31_Current_value(self):
        # self.phy_0.TPSM831D31_VoltageSet(0x1, 'CHA', volt)  # VDD
        # self.phy_0.TPSM831D31_VoltageSet(0x1, 'CHB', volt)  # IOVDD
        self.phy_0.TPSM831D31_CurrentRead(0x2, "CHA")  # D0_AVDD_V2
        self.phy_0.TPSM831D31_CurrentRead(0x2, "CHB")  # D0_AVDD12_V2
        self.phy_0.TPSM831D31_CurrentRead(0x4, "CHA")  # D0_AVDD_V1_D1_V1
        self.phy_0.TPSM831D31_CurrentRead(0x4, "CHB")  # D0_AVDD12_V1_D1_V1
        self.phy_0.TPSM831D31_CurrentRead(0x8, "CHA")  # D1_AVDD_V2_D2_V1
        self.phy_0.TPSM831D31_CurrentRead(0x8, "CHB")  # D1_AVDD12_V2_D2_V1
        self.phy_0.TPSM831D31_CurrentRead(0x10, "CHA")  # D2_AVDD_V2
        self.phy_0.TPSM831D31_CurrentRead(0x10, "CHB")  # D2_AVDD12_V2

    """Instrument"""

    def Thermal_OnOff(self, event):
        if self.ThermalOn_OFF.Value:
            self.ThermalOn_OFF.Label = "Thermal On"
        else:
            self.ThermalOn_OFF.Label = "Thermal OFF"

    def datalog1_visa_chk(self):
        instrument_ID = self.visa.IDN(visa=self.datalog1_visa_wx.Value)
        if instrument_ID.find("34970") != -1:
            self.bgColor = wx.GREEN
            visa_status = "Link_PASS"
        else:
            self.bgColor = wx.RED
            visa_status = "Link_FAIL"
        self.datalog1_visa_status_wx.Value = visa_status
        self.datalog1_visa_status_wx.SetBackgroundColour(self.bgColor)

    def datalog2_visa_chk(self):
        instrument_ID = self.visa.IDN(visa=self.datalog2_visa_wx.Value)
        if instrument_ID.find("34970") != -1:
            self.bgColor = wx.GREEN
            visa_status = "Link_PASS"
        else:
            self.bgColor = wx.RED
            visa_status = "Link_FAIL"
        self.datalog2_visa_status_wx.Value = visa_status
        self.datalog2_visa_status_wx.SetBackgroundColour(self.bgColor)

    def datalog1_visa_even(self, even):
        self.datalog1_visa_chk()

    def datalog2_visa_even(self, even):
        self.datalog2_visa_chk()

    def TA5000A_visa_even(self, even):
        instrument_ID = self.visa.IDN(visa=self.TA5000A_visa_wx.Value)
        # print(instrument_ID)
        if instrument_ID.find("5000") != -1:
            self.bgColor = wx.GREEN
            visa_status = "Link_PASS"
        else:
            self.bgColor = wx.RED
            visa_status = "Link_FAIL"
        self.TA5000A_visa_status_wx.Value = visa_status
        self.TA5000A_visa_status_wx.SetBackgroundColour(self.bgColor)

    def TA5000A_set(self):
        print("Setup Temperature Value")

    def power_info_even(self, even):
        self.power_channel_set()

    def power1_visa_even(self, even):
        self.power_channel_set()

    def power2_visa_even(self, even):
        self.power_channel_set()

    def power3_visa_even(self, even):
        self.power_channel_set()

    def power4_visa_even(self, even):
        self.power_channel_set()

    def power5_visa_even(self, even):
        self.power_channel_set()

    def power_channel_set(self, **kargs):
        skip = kargs.get("skip", 0)

        instrument_ID = self.visa.IDN(visa=self.power1_visa_wx.Value)
        if instrument_ID.find("36") != -1:
            visa_status = "Link_PASS"
            self.bgColor = wx.GREEN
        else:
            visa_status = "Link_FAIL"
            self.bgColor = wx.RED
        self.power1_visa_status_wx.Value = visa_status
        self.power1_visa_status_wx.SetBackgroundColour(self.bgColor)

        instrument_ID = self.visa.IDN(visa=self.power2_visa_wx.Value)
        if instrument_ID.find("36") != -1:
            visa_status = "Link_PASS"
            self.bgColor = wx.GREEN
        else:
            visa_status = "Link_FAIL"
            self.bgColor = wx.RED
        self.power2_visa_status_wx.Value = visa_status
        self.power2_visa_status_wx.SetBackgroundColour(self.bgColor)

        instrument_ID = self.visa.IDN(visa=self.power3_visa_wx.Value)
        if instrument_ID.find("36") != -1:
            visa_status = "Link_PASS"
            self.bgColor = wx.GREEN
        else:
            visa_status = "Link_FAIL"
            self.bgColor = wx.RED
        self.power3_visa_status_wx.Value = visa_status
        self.power3_visa_status_wx.SetBackgroundColour(self.bgColor)

        instrument_ID = self.visa.IDN(visa=self.power4_visa_wx.Value)
        if instrument_ID.find("36") != -1:
            visa_status = "Link_PASS"
            self.bgColor = wx.GREEN
        else:
            visa_status = "Link_FAIL"
            self.bgColor = wx.RED
        self.power4_visa_status_wx.Value = visa_status
        self.power4_visa_status_wx.SetBackgroundColour(self.bgColor)

        instrument_ID = self.visa.IDN(visa=self.power5_visa_wx.Value)
        if instrument_ID.find("36") != -1:
            visa_status = "Link_PASS"
            self.bgColor = wx.GREEN
        else:
            visa_status = "Link_FAIL"
            self.bgColor = wx.RED
        self.power5_visa_status_wx.Value = visa_status
        self.power5_visa_status_wx.SetBackgroundColour(self.bgColor)

        instrument_ID = self.visa.IDN(visa=self.power_cycle_visa.Value)
        if instrument_ID.find("36") != -1:
            visa_status = "Link_PASS"
            self.bgColor = wx.GREEN
        else:
            visa_status = "Link_FAIL"
            self.bgColor = wx.RED
        self.power_cycle_visa_status.Value = visa_status
        self.power_cycle_visa_status.SetBackgroundColour(self.bgColor)

    def datalog_meas_even(self, even):
        Voltage = self.visa.Keysight_DataLog_793()
        # datalog ch 101 : D0_AVDD12_V2
        # datalog ch 102 : D0_AVDD_V2
        # datalog ch 103 : D0_AVDD12_V1_D1_V1
        # datalog ch 104 : IOVDD
        # datalog ch 105 : VDD
        # datalog ch 106 : D1_AVDD_V2_D2_V1
        # datalog ch 107 : D1_AVDD12_V2_D2_V1
        # datalog ch 108 : D2_AVDD_V2
        # datalog ch 109 : D2_AVDD12_V2
        # datalog ch 110 : D0_AVDD_V1_D1_V1
        self.VDD_V_SENSE.Value = str(Voltage[4])
        self.IOVDD_SENSE.Value = str(Voltage[3])
        self.D0_AVDD_V2_SENSE.Value = str(Voltage[1])
        self.D0_AVDD12_V2_SENSE.Value = str(Voltage[0])
        self.D0_AVDD_V1_D1_V1_SENSE.Value = str(Voltage[9])
        self.D0_AVDD12_V1_D1_V1_SENSE.Value = str(Voltage[2])
        self.D1_AVDD_V2_D2_V1_SENSE.Value = str(Voltage[5])
        self.D1_AVDD12_V2_D2_V1_SENSE.Value = str(Voltage[6])
        self.D2_AVDD_V2_SENSE.Value = str(Voltage[7])
        self.D2_AVDD12_V2_SENSE.Value = str(Voltage[8])

    def meter_info_even(self, even):
        if self.meter_info_wx.Validate() == True:
            self.meter_info_wx.Label = "Checking VISA"
            self.datalog1_visa_even("none")
            self.datalog2_visa_even("none")
            self.meter_info_wx.Label = "Press Check VISA Bus"
        else:
            self.meter_info_wx.Label = "Press Check VISA Bus"

    """Test Report"""

    def xlsx_Report(self, **kargs):
        if self.Thermal_die_en.Value == True:
            Data_log = self.visa.Keysight_DataLog_793_101_104(
                visa="USB0::0x2A8D::0x5101::MY58014090::0::INSTR"
            )
            v1 = Data_log[0]
            v2 = Data_log[1]
            v3 = Data_log[2]
            self.die0_THM = self.run_0.THM_Value(v1)
            self.die1_THM = self.run_0.THM_Value(v2)
            self.die2_THM = self.run_0.THM_Value(v3)
            print(f"Die0_Thermal Temp Value={self.die0_THM} Degree C")
            print(f"Die1_Thermal Temp Value={self.die1_THM} Degree C")
            print(f"Die2_Thermal Temp Value={self.die2_THM} Degree C")
            result = (
                [self.TestItem]
                + [self.chip_version]
                + self.test_condition
                + self.Test_Info_list
                + ["Hyperlink_Log"]
                + [""]
                + self.slice_result
                + [self.die0_THM]
                + [self.die1_THM]
                + [self.die2_THM]
            )
        else:
            result = (
                [self.TestItem]
                + [self.chip_version]
                + self.test_condition
                + self.Test_Info_list
                + ["Hyperlink_Log"]
                + self.slice_result
            )
        Start_row = self.xlsx_write_result(result, self.log_path)

        # if self.build_eye_graph_wx.CurrentSelection==0:
        #     for L in range(1):
        #         # print(f'\n< Edit Eye Diagram Test Graph{L} >', flush=True)
        #         print('< Edit Eye Diagram Test Graph >', flush=True)
        #         if self.eye_graph_en != 0:
        #             result_arr = [self.s0_eye_result,self.s1_eye_result,self.s2_eye_result,self.s3_eye_result]
        #             name_arr = ['Eye1','Eye2','Eye3','Eye4']
        #             graph_arr = ['Graph_Eye1','Graph_Eye2','Graph_Eye3','Graph_Eye4']
        #
        #             for r in range (len(result_arr)):
        #                 result = result_arr[r]
        #                 name = name_arr[r]
        #                 textfile = open(f"TestTools/{graph_arr[r]}.txt", "w")
        #                 for i in range(len(result)):
        #                     content = str(result[i].flatten().tolist())
        #                     content = re.sub(",", "", content)
        #                     content = content.replace("[", "")
        #                     content = content.replace("]", "")
        #                     content = f'{content}\n'
        #                     textfile.write(content)
        #
        #                 # creat folder
        #                 folder_name = ((self.save_log).split('.txt'))[0]
        #                 if r ==0 and L==0:
        #                     os.mkdir(folder_name)
        #                 textfile.write(f'{self.tools_path}/{folder_name}')
        #
        #                 # eye daigram name
        #                 # daigram_name = f',/{self.Log_Folder_path}_{name}'
        #                 daigram_name = f',/{name}'
        #                 textfile.write(daigram_name)
        #                 textfile.close()
        #             os.system("Graph.exe")
        #             time.sleep(5)
        #             command = "taskkill /f /t /im Graph.exe"
        #             os.system(command)
        #
        #             self.picture_merge(folder_name)
        # else:
        #     print('(Skip This Function)')
        #     pass

    def xlsx_write_result(self, Test_Result, Hyperlink_path):
        Excel_Path = load_workbook(filename=self.xls_report_path)
        Select_sheet = Excel_Path["Test Result"]

        for prcs in psutil.process_iter():
            if (
                str(prcs.name).find("EXCEL.EXE") != -1
            ):  # 印出所有正在執行的應用程式 ( 從中觀察 pid )
                print("Close Excel Tools")
                self.event_window()
                break
            else:
                pass

        for i in range(99999999999):
            Start_row = Select_sheet.cell(row=i + 5, column=1).value
            # print(Start_row)
            if Start_row == None:
                break
            else:
                pass
        Start_row_Num = i + 5
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i in range(len(Test_Result)):
            Select_sheet.cell(row=Start_row_Num, column=i + 1).alignment = align
            Select_sheet.cell(row=Start_row_Num, column=i + 1).border = border
            if "FAIL" in Test_Result or "abp_failed" in Test_Result:
                font = Font(
                    "Arial",
                    size=10,
                    bold=True,
                    italic=False,
                    strike=False,
                    color="ff0000",
                )
            else:
                font = Font(
                    "Arial",
                    size=10,
                    bold=False,
                    italic=False,
                    strike=False,
                    color="000000",
                )
            if Test_Result[i] == "Hyperlink_Log":
                font = Font(
                    "Arial",
                    size=10,
                    bold=False,
                    italic=False,
                    strike=False,
                    color="0000ff",
                )
                Select_sheet.cell(
                    row=Start_row_Num, column=i + 1
                ).hyperlink = Hyperlink_path
                Select_sheet.cell(row=Start_row_Num, column=i + 1).style = "Hyperlink"
                Select_sheet.cell(
                    row=Start_row_Num, column=i + 1
                ).value = "HYPERLINK LINK"
                Select_sheet.cell(row=Start_row_Num, column=i + 1).alignment = align
                Select_sheet.cell(row=Start_row_Num, column=i + 1).border = border
                Select_sheet.cell(row=Start_row_Num, column=i + 1).font = font
                # print('Hyperlink_Log')
            else:
                Select_sheet.cell(row=Start_row_Num, column=i + 1).font = font
                Select_sheet.cell(row=Start_row_Num, column=i + 1).value = Test_Result[
                    i
                ]
        Excel_Path.save(self.xls_report_path)
        Excel_Path.close()
        if self.eye_graph_en != 0:
            self.xls_graph_result(
                row=Start_row_Num, col=i, Hyperlink_path=Hyperlink_path
            )

        self.current = Start_row_Num
        return Start_row_Num

    def xlsx_log_file_Name(self, result, sheet_name, test_num):
        Excel_Path = load_workbook(filename=self.xls_report_path)
        Select_sheet = Excel_Path[sheet_name]

        for prcs in psutil.process_iter():
            if (
                str(prcs.name).find("EXCEL.EXE") != -1
            ):  # 印出所有正在執行的應用程式 ( 從中觀察 pid )
                print("Close Excel Tools")
                self.event_window()
                break
            else:
                pass

        for i in range(99999999999):
            Start_row = Select_sheet.cell(row=i + 4, column=8).value
            # print(Start_row)
            if Start_row == None:
                break
            else:
                pass

        col = 1
        Start_row_Num = i + 4
        for n in range(len(result)):
            # print(f'(SW) Result {n + 1} / {len(result)}', flush=True)

            # 'Die0V_Slice0_Vref=Center : w_s=23 , w(%)=71.8% , 0000000000000111111111000000000000000000000001111111100000000000(BIN / Pi_Step_Min(Zero) TO Pi_Step_Max)'
            aa = result[n].split("_")
            die_str = (result[n].split("_"))[0]
            slice_str = (result[n].split("_"))[1]
            vref_str = (((result[n].split(":"))[0]).split("_"))[2]
            w_s = ((result[n].split(":"))[1]).split(",")[0]
            pa = ((result[n].split(":"))[1]).split(",")[1]
            file_Name = (
                [self.Log_Folder_path]
                + [self.TestDataRate]
                + [die_str]
                + [slice_str]
                + [vref_str]
                + [w_s]
                + [pa]
                + [self.test_lane]
                + [self.sw_bist]
            )
            info_len = len(file_Name)
            for D in range(info_len):
                border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"),
                )
                align = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                font = Font("Arial", size=10, bold=True, italic=False, strike=False)
                if die_str.find("-") != -1:
                    Select_sheet.cell(
                        row=Start_row_Num + n, column=col + D
                    ).alignment = align
                    Select_sheet.cell(
                        row=Start_row_Num + n, column=col + D
                    ).border = border
                    Select_sheet.cell(row=Start_row_Num + n, column=col + D).font = font
                    Select_sheet.cell(
                        row=Start_row_Num + n, column=col + D
                    ).value = "--"
                else:
                    if D == 0:
                        Hyperlink_path = (
                            "Test Report Log/" + self.Log_Folder_path + ".txt"
                        )
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).hyperlink = Hyperlink_path
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).style = "Hyperlink"
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).alignment = align
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).border = border
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).font = font
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).value = file_Name[D]
                    else:
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).alignment = align
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).border = border
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).font = font
                        Select_sheet.cell(
                            row=Start_row_Num + n, column=col + D
                        ).value = file_Name[D]

            bin = (",".join(((result[n].split(" , ")[2]).split("("))[0])).split(",")
            for i in range(len(bin)):
                if bin[i] == "1" or bin[i] == "V" or bin[i] == "E":
                    fill = PatternFill(
                        start_color="FF0000", end_color="FF0000", fill_type="solid"
                    )
                    bin_log = bin[i]
                elif bin[i].find("X") != -1:
                    fill = PatternFill(
                        start_color="7F7F7F", end_color="7F7F7F", fill_type="solid"
                    )
                    bin_log = "x"
                elif bin[i].find("-") != -1:
                    fill = PatternFill(
                        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                    )
                    bin_log = "-"
                else:
                    fill = PatternFill(
                        start_color="00FF00", end_color="00FF00", fill_type="solid"
                    )
                    bin_log = bin[i]
                font = Font("Arial", size=10, bold=False, italic=False, strike=False)
                Select_sheet.cell(
                    row=Start_row_Num + n, column=i + info_len + 1
                ).alignment = align
                Select_sheet.cell(
                    row=Start_row_Num + n, column=i + info_len + 1
                ).border = border
                Select_sheet.cell(
                    row=Start_row_Num + n, column=i + info_len + 1
                ).font = font
                Select_sheet.cell(
                    row=Start_row_Num + n, column=i + info_len + 1
                ).fill = fill
                Select_sheet.cell(
                    row=Start_row_Num + n, column=i + info_len + 1
                ).value = bin_log

            # for i in range(71):
            #     Select_sheet.cell(row=Start_row_Num+n+1, column=i + 1).alignment = align
            #     Select_sheet.cell(row=Start_row_Num+n+1, column=i + 1).border = border
            #     font = Font(u'Arial', size=10, bold=True, italic=False, strike=False)
            #     Select_sheet.cell(row=Start_row_Num+n+1, column=i + 1).font = font
            #     Select_sheet.cell(row=Start_row_Num+n+1, column=i + 1).value = ' --'
        Excel_Path.save(self.xls_report_path)
        Excel_Path.close()

        return Start_row_Num

    def xlsx_log_link(self):
        current_num = f"FY{self.current}"
        target_num = f"A{self.target_num}"

        xlsx_path = load_workbook(filename=self.xls_report_path)
        ws_current = xlsx_path["Test Result"]
        ws_target = xlsx_path["Eye Diagram Width"]
        ws_target[target_num] = self.Log_Folder_path
        ws_current[current_num] = "Eye Diagram Graph"
        ws_current[current_num].hyperlink = f"#'Eye Diagram Width'!{target_num}"
        ws_current[current_num].style = "Hyperlink"
        xlsx_path.save(self.xls_report_path)

        # xlsx_path = load_workbook(filename=self.xls_report_path)
        # ws_current = xlsx_path['Test Result']
        # ws_target = xlsx_path['Eye Diagram']
        # ws_target['B2'] = 'Eye log'
        # ws_current['B16'] = 'eye link'
        # ws_current['B16'].hyperlink = "#'Eye Diagram'!B2"
        # ws_current['B16'].style = 'Hyperlink'
        # xlsx_path.save(self.xls_report_path)

    def xls_graph_result(self, **kargs):
        row = kargs.get("row", 0)
        col = kargs.get("col", 0)
        Hyperlink_path = kargs.get("Hyperlink_path", "NA")

        Excel_Path = load_workbook(filename=self.xls_report_path)
        Select_sheet = Excel_Path["Test Result"]
        font = Font(
            "Arial", size=10, bold=False, italic=False, strike=False, color="0000ff"
        )
        HYPERLINK = (
            '=HYPERLINK("'
            + Hyperlink_path[:-3]
            + "png"
            + '",'
            + '"'
            + "Hyperlink_Graph"
            + '"'
            + ")"
        )
        Select_sheet.cell(row=row, column=col + 1).value = HYPERLINK
        Select_sheet.cell(row=row, column=col + 1).font = font
        Excel_Path.save(self.xls_report_path)
        Excel_Path.close()

    def xlsx_cut(self, sheet_name):
        Excel_Path = load_workbook(filename=self.xls_report_path)
        Select_sheet = Excel_Path[sheet_name]

        for prcs in psutil.process_iter():
            if (
                str(prcs.name).find("EXCEL.EXE") != -1
            ):  # 印出所有正在執行的應用程式 ( 從中觀察 pid )
                print("Close Excel Tools")
                self.event_window()
                break
            else:
                pass

        for i in range(99999999999):
            Start_row = Select_sheet.cell(row=i + 4, column=6).value
            # print(Start_row)
            if Start_row == None:
                break
            else:
                pass
        Start_row_Num = i + 4
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        Start_row_Num = Start_row_Num
        for i in range(135):
            Select_sheet.cell(row=Start_row_Num, column=i + 1).alignment = align
            Select_sheet.cell(row=Start_row_Num, column=i + 1).border = border
            font = Font("Arial", size=10, bold=True, italic=False, strike=False)
            Select_sheet.cell(row=Start_row_Num, column=i + 1).font = font
            Select_sheet.cell(row=Start_row_Num, column=i + 1).fill = fill
            Select_sheet.cell(row=Start_row_Num, column=i + 1).value = " --"
        Excel_Path.save(self.xls_report_path)
        Excel_Path.close()

    def xlsx_reg_compare(self, Test_Result):
        Excel_Path = load_workbook(filename=self.xls_report_path)
        Select_sheet = Excel_Path["Register_Compare"]

        for prcs in psutil.process_iter():
            if (
                str(prcs.name).find("EXCEL.EXE") != -1
            ):  # 印出所有正在執行的應用程式 ( 從中觀察 pid )
                print("Close Excel Tools")
                self.event_window()
                break
            else:
                pass

        for i in range(99999999999):
            Start_row = Select_sheet.cell(row=i + 4, column=6).value
            # print(Start_row)
            if Start_row == None:
                break
            else:
                pass
        Start_row_Num = i + 4
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        Start_row_Num = Start_row_Num
        for i in range(len(Test_Result)):
            result_list = (Test_Result[i]).split(",")
            for r in range(len(result_list)):
                if result_list[7] == "Failed":
                    font = Font(
                        "Arial",
                        size=10,
                        bold=True,
                        italic=False,
                        strike=False,
                        color="ff0000",
                    )
                else:
                    font = Font(
                        "Arial",
                        size=10,
                        bold=False,
                        italic=False,
                        strike=False,
                        color="000000",
                    )
                Select_sheet.cell(row=Start_row_Num + i, column=r + 1).alignment = align
                Select_sheet.cell(row=Start_row_Num + i, column=r + 1).border = border
                Select_sheet.cell(row=Start_row_Num + i, column=r + 1).font = font
                Select_sheet.cell(
                    row=Start_row_Num + i, column=r + 1
                ).value = result_list[r]

        Excel_Path.save(self.xls_report_path)
        Excel_Path.close()

    """Graph"""

    def picture_merge(self, folder_path, **kwargs):
        row_width = 6  # one col have 6 picture
        pic_width = 530
        pic_high = 480
        eye_pic_arr = next(os.walk(folder_path))[2]
        pic_num = len(eye_pic_arr)
        col = int(pic_num / row_width)

        from PIL import Image

        bg = Image.new(
            "RGB", (pic_width * col, pic_high * row_width), "#000000"
        )  # pic_width x pic_high col / row
        for i in range(1, pic_num + 1):
            name = f"{folder_path}/{eye_pic_arr[i - 1]}"
            # print(name)
            img = Image.open(name)  # 開啟圖片
            img = img.resize((pic_width, pic_high))  # set graph size 300x400
            y = (i - 1) % row_width  # 根據開啟的順序，決定 x 座標
            x = (i - 1) // row_width  # 根據開啟的順序，決定 y 座標 ( // 為快速取整數 )
            bg.paste(img, (x * pic_width, y * pic_high))

        path = os.getcwd() + "/TestTools/TestTools_GUI.png"
        bg.save(path)
        path = "Test Report/" + ((self.log_path).split("txt"))[0] + "png"
        bg.save(path)

    """eye scan"""

    def eye_scan_even(self, event):
        self.m_richText1.Clear()
        print("\n< Eye Scan Test Start > ")
        if self.eye_scan_en == 0:
            self.TestItem_Now2_wx.SetBackgroundColour("#FF0000")
            self.TestItem_Now2_wx.Value = f'Need to run  "Hardware Training Test"'
            self.eye_scan_wx.Value = False
        else:
            self.eye_scan_en = 0
            self.info_window_wx.Selection = 1
            self.eye_scan_wx.Value = True
            self.eye_scan_wx.Label = "Eye Scan Now"
            self.TestItem_Now2_wx.SetBackgroundColour("#000000")
            self.TestItem_Now2_wx.Value = "Eye Scan Test"
            self.TestItem_Now_wx.Value = "( Test Condition )"

            str = self.scan_cycle_wx.GetValue()
            scan_event = str.upper()
            if scan_event == "NA":
                for w in range(999999 * 999999 * 999999):
                    # self.HW_Training_init()
                    if w == 0:
                        self.eye_scan_en == 0
                    else:
                        self.eye_scan_en == 1
                    self.TestItem_Now_wx.Value = (
                        f"Auto Eye Scan - {w + 1} (On Going !! )"
                    )
                    self.Step_count.Value = w
                    self.Step_count.Range = 1
                    self.m_textCtrl9.Value = f"Test Cycle {w} (Without Limits) "
                    if w == 0:
                        os.popen("TestTools\\event")
                    # print(w, flush=True)

                    self.HW_Training_init()
                    self.m_richText1.Clear()

                    get_scan = self.get_win()
                    if get_scan == False:
                        break
                    else:
                        pass
            elif scan_event == "1":
                self.eye_scan_en == 0
                self.Step_count.Value = 1
                self.Step_count.Range = 1
                self.m_textCtrl9.Value = f"Test Cycle 1 of 1 "
                self.TestItem_Now_wx.Value = f"Single Eye Scan (On Going !! )"
                self.HW_Training_init()
            else:
                scan_num = int(self.scan_cycle_wx.GetValue())
                for s in range(scan_num):
                    if s == 0:
                        self.eye_scan_en == 0
                    else:
                        self.eye_scan_en == 1
                    self.TestItem_Now_wx.Value = (
                        f"Auto Eye Scan - {s + 1}  (On Going !! )"
                    )
                    self.Step_count.Value = s
                    self.Step_count.Range = scan_num
                    self.m_textCtrl9.Value = f"Test Cycle {s} of {scan_num} "
                    if s == 0:
                        os.popen("TestTools\\event")
                    # print(w, flush=True)

                    self.HW_Training_init()
                    self.m_richText1.Clear()

                    get_scan = self.get_win()
                    if get_scan == False:
                        break
                    else:
                        pass
            self.Step_count.Value = self.Step_count.Value + 1
            self.TestItem_Now_wx.Value = (
                "( Create Receiver Eye Scan Graph and Result App )"
            )
            command = "taskkill /f /t /im event.exe"
            os.system(command)
            self.eye_scan_wx.Label = "Press Run Eye Scan"
            print("\nEye Scan Test Done !! ")

            self.eye_scan_window()
            self.TestItem_Now2_wx.Value = "( Test Function )"
            self.TestItem_Now_wx.Value = "( Test Condition )"
            # self.m_textCtrl9.Value = 'Test Done'
            self.eye_scan_wx.Value = False

    def retry_show_fail_count_event(self, event):
        self.retry_show_count.Label = "Update Fail Count Now"
        self.retry_show_count.Value = True
        self.eye_scan_window()
        self.retry_show_count.Label = "Press Update Fail Count"
        self.retry_show_count.Value = False

    def eye_scan_window(self):
        # for g in range(32):
        #     print(self.eye_result[g])
        eye_scan_type_num = self.eye_scan_type.GetSelection()
        if eye_scan_type_num == 0:
            self.eye_result_show = self.eye_result_s0
        elif eye_scan_type_num == 1:
            self.eye_result_show = self.eye_result_s1
        elif eye_scan_type_num == 2:
            self.eye_result_show = self.eye_result_s2
        elif eye_scan_type_num == 3:
            self.eye_result_show = self.eye_result_s3
        else:
            pass

        # eye diagram test result value
        vref_c = (self.eye_result_show)[15]  # cneter vref
        for r in range(len(vref_c)):
            if vref_c[r] == 0:
                break
            else:
                pass
        self.w_l = r + 1  # phase pass left limit
        for r in range(len(vref_c)):
            if vref_c[r + self.w_l] != 0:
                break
            else:
                pass
        self.w_r = r + self.w_l  # phase pass left limit
        self.scan_info_win = (
            f"Eye Width : Eye Width Left = {self.w_l} , Eye Width Right = {self.w_r}"
        )

        rbvs_v = []
        for h in range(self.vref_num):
            vef_list = (self.eye_result_show)[h]
            rbv = vef_list[32]  # any vref ' s center phase value
            rbvs_v.append(rbv)
        # print(rbvs_v)
        for r in range(len(rbvs_v)):
            if rbvs_v[r] == 0:
                break
            else:
                pass
        self.vref_top = r + 1  # phase pass left limit
        for r in range(len(rbvs_v)):
            if rbvs_v[r + self.vref_top] != 0:
                break
            else:
                pass
        self.vref_bottom = r + self.vref_top  # phase pass left limit
        self.scan_info_high = f"Eye High : Eye High Top = {self.vref_top} , Eye High Bottom = {self.vref_bottom}"

        max_value = (self.eye_result_show[0])[0]

        root = tk.Tk()
        fontStyle2 = tkFont.Font(family="Lucida Bright", size=8)
        root.title("Receiver Eye Sacn Graph")

        # print(self.eye_result)
        verf_len = len(self.eye_result_show)
        phase_len = len(self.eye_result_show[0])
        step = 20
        canvas = tk.Canvas(root, width=phase_len * step, height=verf_len * step)
        root.geometry(f"{phase_len * step - 30}x{verf_len * step - 50}+0+0")

        ysb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
        xsb = tk.Scrollbar(root, orient="horizontal", command=canvas.xview)
        canvas.configure(xscrollcommand=xsb.set, yscrollcommand=ysb.set)
        root.grid_rowconfigure(0, weight=1)
        root.grid_columnconfigure(0, weight=1)
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")

        group1 = tk.LabelFrame(
            root,
            text=f"< {self.eye_scan_type.GetValue()} Eye Scan Test Result >",
            font=fontStyle2,
        )
        group2 = tk.LabelFrame(root, text="< Color Grade >", font=fontStyle2)

        center_color_TF = fail_color_1_TF = fail_color_max_TF = 0
        fail_color_r1_TF = fail_color_r2_TF = fail_color_r3_TF = fail_color_r4_TF = (
            fail_color_r5_TF
        ) = 0
        fail_color_r6_TF = fail_color_r7_TF = fail_color_r8_TF = fail_color_r9_TF = (
            fail_color_r10_TF
        ) = 0
        for x in range(0, verf_len, 1):
            for y in range(0, phase_len, 1):
                value = self.eye_result_show[x]
                # print(value)
                value = value[y]
                if value == 0:
                    if x == 15 and y == 32:
                        center_color = "#006600"
                    else:
                        center_color = "#00cc00"
                    center_color_TF = 1

                if max_value == 1:
                    fail_color = fail_color_1 = "red"
                    fail_color_1_TF = 1
                elif value == max_value:
                    fail_color = fail_color_max = "#2F0000"
                    fail_color_max_TF = 1
                else:
                    if value == 1:
                        fail_color = fail_color_1 = "#FFB5B5"
                        fail_color_1_TF = 1
                    else:
                        color_list = (
                            "#FF7575",
                            "#FF5151",
                            "#FF2D2D",
                            "#FF0000",
                            "#EA0000",
                            "#CE0000",
                            "#930000",
                            "#750000",
                            "#600000",
                            "#4D0000",
                        )
                        if max_value <= 10:
                            grade = 1
                        else:
                            grade = max_value // len(color_list)
                        if 2 < value <= grade:
                            fail_color_r1 = fail_color = color_list[0]
                            fail_color_r1_TF = 1
                        elif grade * 1 < value <= grade * 2:
                            fail_color_r2 = fail_color = color_list[1]
                            fail_color_r2_TF = 1
                        elif grade * 2 < value <= grade * 3:
                            fail_color_r3 = fail_color = color_list[2]
                            fail_color_r3_TF = 1
                        elif grade * 3 < value <= grade * 4:
                            fail_color_r4 = fail_color = color_list[3]
                            fail_color_r4_TF = 1
                        elif grade * 4 < value <= grade * 5:
                            fail_color_r5 = fail_color = color_list[4]
                            fail_color_r5_TF = 1
                        elif grade * 5 < value <= grade * 6:
                            fail_color_r6 = fail_color = color_list[5]
                            fail_color_r6_TF = 1
                        elif grade * 6 < value <= grade * 7:
                            fail_color_r7 = fail_color = color_list[6]
                            fail_color_r7_TF = 1
                        elif grade * 7 < value <= grade * 8:
                            fail_color_r8 = fail_color = color_list[7]
                            fail_color_r8_TF = 1
                        elif grade * 8 < value <= grade * 9:
                            fail_color_r9 = fail_color = color_list[8]
                            fail_color_r9_TF = 1
                        elif grade * 9 < value <= max_value - 1:
                            fail_color_r10 = fail_color = color_list[9]
                            fail_color_r10_TF = 1
                        else:
                            pass

                color = center_color if value == 0 else fail_color
                scan = 15
                if (
                    int(self.seach_min.Value) <= value
                    and int(self.seach_max.Value) >= value
                ):
                    fail_count = value
                else:
                    fail_count = ""
                canvas.create_rectangle(
                    scan * y,
                    scan * (x + 1),
                    scan * (y + 1),
                    scan * 2 + (scan * x),
                    fill=color,
                    outline="#fff",
                )
                canvas.create_text(
                    scan * 0.5 + (scan * y),
                    scan * 1.5 + (scan * x),
                    text=fail_count,
                    fill="white",
                    font=("Lucida Bright", 6),
                )

        # label
        vref_s = tk.Label(root, text="(Vref 0)", font=fontStyle2)
        vref_d = tk.Label(root, text="(Vref 31)", font=fontStyle2)
        phase_s = tk.Label(root, text="(Phase -31)", font=fontStyle2)
        phase_d = tk.Label(root, text="(Phase 31)", font=fontStyle2)
        tk.Label(group1, text=self.scan_info_win, font=fontStyle2).grid(
            column=0, row=0, sticky=tk.W
        )
        tk.Label(group1, text=self.scan_info_high, font=fontStyle2).grid(
            column=0, row=1, sticky=tk.W
        )

        tk.Label(group2, text="Center Phase Pass", fg="#006600", font=fontStyle2).grid(
            column=0, row=0, sticky=tk.W
        )
        if center_color_TF:
            tk.Label(group2, text="Phase Pass", fg=center_color, font=fontStyle2).grid(
                column=0, row=1, sticky=tk.W
            )
        if fail_color_1_TF:
            tk.Label(
                group2, text="Phase Fail 1 Time", fg=fail_color_1, font=fontStyle2
            ).grid(column=0, row=2, sticky=tk.W)
        if fail_color_max_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {max_value} Times",
                fg=fail_color_max,
                font=fontStyle2,
            ).grid(column=0, row=3, sticky=tk.W)
        if fail_color_r1_TF:
            tk.Label(
                group2,
                text=f"Phase Fail 2 To {grade} Times",
                fg=fail_color_r1,
                font=fontStyle2,
            ).grid(column=0, row=4, sticky=tk.W)
        if fail_color_r2_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade} To {grade * 2} Times",
                fg=fail_color_r2,
                font=fontStyle2,
            ).grid(column=0, row=5, sticky=tk.W)
        if fail_color_r3_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 2} To {grade * 3} Times",
                fg=fail_color_r3,
                font=fontStyle2,
            ).grid(column=0, row=6, sticky=tk.W)
        if fail_color_r4_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 3} To {grade * 4} Times",
                fg=fail_color_r4,
                font=fontStyle2,
            ).grid(column=0, row=7, sticky=tk.W)
        if fail_color_r5_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 4} To {grade * 5} Times",
                fg=fail_color_r5,
                font=fontStyle2,
            ).grid(column=0, row=8, sticky=tk.W)
        if fail_color_r6_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 5} To {grade * 6} Times",
                fg=fail_color_r6,
                font=fontStyle2,
            ).grid(column=0, row=9, sticky=tk.W)
        if fail_color_r7_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 6} To {grade * 7} Times",
                fg=fail_color_r7,
                font=fontStyle2,
            ).grid(column=0, row=10, sticky=tk.W)
        if fail_color_r8_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 7} To {grade * 8} Times",
                fg=fail_color_r8,
                font=fontStyle2,
            ).grid(column=0, row=11, sticky=tk.W)
        if fail_color_r9_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 8} To {grade * 9} Times",
                fg=fail_color_r9,
                font=fontStyle2,
            ).grid(column=0, row=12, sticky=tk.W)
        if fail_color_r10_TF:
            tk.Label(
                group2,
                text=f"Phase Fail {grade * 9} To {max_value - 1} Times",
                fg=fail_color_r10,
                font=fontStyle2,
            ).grid(column=0, row=13, sticky=tk.W)

        # Buttom setup
        autoButton = tk.Button(
            root, command=root.destroy, text=" Close Window ", font=fontStyle2
        )
        group1.grid()
        group2.grid()
        add = 10
        group1.place(x=65 + add, y=0)
        group2.place(x=1030 + add, y=55)
        vref_s.place(x=980 + add, y=40)
        phase_s.place(x=0, y=verf_len * step - 120)
        phase_d.place(x=1030 + add, y=verf_len * step - 120)
        canvas.place(x=65 + add, y=45)
        vref_d.place(x=975 + add, y=verf_len * step - 100)
        autoButton.place(x=65 + add, y=verf_len * step - 95)

        root.mainloop()

    def merge_one_worst(self):
        self.s0_eye_result = eye_sum_s0 = np.array(self.slice_result[0])
        self.s1_eye_result = eye_sum_s1 = np.array(self.slice_result[1])
        self.s2_eye_result = eye_sum_s2 = np.array(self.slice_result[2])
        self.s3_eye_result = eye_sum_s3 = np.array(self.slice_result[3])
        if self.eye_scan_en == 1:
            eye_sum_s0 = self.eye_result_s0 + eye_sum_s0
            eye_sum_s1 = self.eye_result_s1 + eye_sum_s1
            eye_sum_s2 = self.eye_result_s2 + eye_sum_s2
            eye_sum_s3 = self.eye_result_s3 + eye_sum_s3
        self.eye_result_s0 = eye_sum_s0
        self.eye_result_s1 = eye_sum_s1
        self.eye_result_s2 = eye_sum_s2
        self.eye_result_s3 = eye_sum_s3

        import gc

        del eye_sum_s0
        gc.collect()
        del eye_sum_s1
        gc.collect()
        del eye_sum_s2
        gc.collect()
        del eye_sum_s3
        gc.collect()

    def print_eye_log(self, result, sheet_name, test_num):
        for n in range(len(result)):
            # 'Die0V_Slice0_Vref=Center : w_s=23 , w(%)=71.8% , 0000000000000111111111000000000000000000000001111111100000000000(BIN / Pi_Step_Min(Zero) TO Pi_Step_Max)'
            die_str = (result[n].split("_"))[0]
            slice_str = (result[n].split("_"))[1]
            vref_str = (((result[n].split(":"))[0]).split("_"))[2]
            w_s = ((result[n].split(":"))[1]).split(",")[0]
            pa = ((result[n].split(":"))[1]).split(",")[1]
            file_Name = (
                [self.Log_Folder_path]
                + [self.TestDataRate]
                + [die_str]
                + [slice_str]
                + [vref_str]
                + [w_s]
                + [pa]
            )
            info_len = len(file_Name)
            if n == 0 or n == 65:
                print(f"2D Eye Diagram_{die_str}_{slice_str}")
            else:
                pass
            bin = (",".join(((result[n].split(" , ")[2]).split("("))[0])).split(",")
            bin_len = int((len(bin)) / 1)
            bin = bin[0:bin_len]
            print(bin)

    """Others"""

    def test_clear_to_org(self):
        self.spec = None
        self.phy_0 = None
        self.run_0 = None

    def project_select(self, event):
        if self.project_json:
            json_ip = self.ip_version + "_D2D"
            if json_ip in self.project_json.keys():
                if len(json_ip.split("_")) >= 2:
                    print("start import py")

                    project_Specialized = getattr(
                        __import__("Specialized"), self.spec_version
                    )
                    project_phy = getattr(__import__("Glink_phy"), self.spec_version)
                    project_run = getattr(__import__("Glink_run"), self.spec_version)
                    project_function = getattr(
                        __import__("Glink_function"), self.spec_version
                    )

                    self.Corner_Version_wx.Clear()
                    for i in self.project_json[json_ip]["Corner_Version_wx_json"]:
                        self.Corner_Version_wx.Append(i)
                    self.Corner_Version_wx.Selection = self.project_json[
                        "i2c_project_speed_mode_test"
                    ][0]

                    self.Function_select_wx.Clear()
                    for i in self.project_json[json_ip]["Function_select_wx_json"]:
                        self.Function_select_wx.Append(i)
                    self.Function_select_wx.Selection = self.project_json[
                        "i2c_project_speed_mode_test"
                    ][1]

                    try:
                        TP_use_get = self.project_json[json_ip]["TP_use"]
                        self.TP_use = 1 if TP_use_get else 0
                        ini_with_rst = self.project_json[json_ip]["ini_with_rst"]
                        self.ini_wo_reset = 0 if ini_with_rst else 1
                    except Exception as e:
                        print(e)
                        self.TP_use = 0
                        self.ini_wo_reset = 1
                    if self.TP_use == 1:
                        print("Set TP sense method")
                    else:
                        print("Set normal sense method (not TP)")
                    try:
                        instrument_visa = self.project_json[json_ip]["instrument_visa"]

                    except Exception as e:
                        print(e)
                        print("instrument information get fail in json file")
                    try:
                        instrument_DCAL_get = self.project_json[json_ip][
                            "instrument_DCAL"
                        ]
                        for instrument_DCAL_list in instrument_DCAL_get:
                            instrument_DCAL = instrument_DCAL_list.split(",")
                            if instrument_DCAL[0] in "B2962A_Current_Setup":
                                self.B2962A_CH1_V21.Value = instrument_DCAL[1]
                                self.B2962A_CH2_V21.Value = instrument_DCAL[2]
                    except Exception as e:
                        print(e)
                        print("instrument dc current limited get fail in json file")

                    self.phy_0 = project_phy(self.i2c, self.jtag, self)
                    self.run_0 = project_run(self.phy_0, self)
                    self.spec = project_Specialized(self.phy_0, self)
                    self.func = project_function(self.phy_0, self)
                else:
                    self.test_clear_to_org()
            else:
                self.test_clear_to_org()
            get_json = True
        else:
            get_json = False
            print(f"project_select get projet_json fail")
        return get_json

    def PASS_FAIL_HW_chk(self):
        buffer = self.m_richText1.Value
        Test_Log = re.sub("\n", "", buffer)
        if Test_Log.find("MBT Failed") != -1:
            return_val = "FAIL"
        else:
            return_val = "PASS"
        return return_val

    def Save_i2cLog(self, **kargs):
        content = kargs.get("log_name", "NA")

        textfile = open("TestTools/i2c_log.txt", "a+")
        textfile.write(content)
        textfile.close()

    def clear_txt(self, file_name, **kargs):
        content = kargs.get("log_name", "NA")

        path = f"TestTools/{file_name}"
        textfile = open(path, "w")
        textfile.write("")
        textfile.close()

    def def_gui(self, **kwargs):
        clear_count = kwargs.get("clear_count", 1)

        color = (0, 0, 0, 0)
        self.TestItem_Now2_wx.SetBackgroundColour(color)
        self.TestItem_Now2_wx.Value = "( Test Function )"
        self.TestItem_Now_wx.Value = "( Test Condition )"
        if clear_count:
            self.Step_count.Value = 0
        self.m_richText1.Clear()
        print("( Test Log )")

    def total_lines(self, path):
        with open(path) as myfile:
            total_lines = sum(1 for line in myfile)
        return total_lines

    def pll_seach_xls_str(self, df, string_s, string_d, **kwargs):
        reg_vai_num = kwargs.get("reg_vai_num", 0)

        for r0 in range(9999):
            buffer1 = df.iloc[r0]
            buffer2 = str(buffer1.iloc[0])
            if buffer2.find(string_s) != -1:
                break
            else:
                pass

        reg_arr = []
        for r in range(9999):
            reg_set = df.iloc[r + r0 + 1]
            reg_name = reg_set.iloc[0]
            offset = reg_set.iloc[1]
            # print(offset,flush=True)
            if str(offset) == "nan" or str(offset) == "":
                pass
            else:
                bit = reg_set.iloc[2]
                value = reg_set.iloc[reg_vai_num]
                edit_log = reg_set.iloc[3]
                ehost = reg_set.iloc[4]
                write_en = reg_set.iloc[5]
                read_en = reg_set.iloc[6]
                die_list = reg_set.iloc[7]
                v_list = reg_set.iloc[8]
                slice_list = reg_set.iloc[9]
                reg = f"{offset}, {bit}, {value}, {edit_log}, {ehost}, {write_en}, {read_en}, {die_list}, {v_list}, {slice_list}"
                reg_arr += [reg]
            if (str(reg_name)).find(string_d) != -1:
                break
            else:
                pass
            # print(f'{reg}\n', flush=True)
        return reg_arr

    def lane_seach_xls_str(self, df, **kwargs):
        reg_col_num = kwargs.get("reg_col_num", 0)

        lane_set_arr = []
        for r in range(70):
            set_list = df.iloc[r + 8]
            lane_set_arr.append(set_list.iloc[reg_col_num])
        self.lane_set_arr = lane_set_arr

    def seach_xls_str(self, df, **kwargs):
        reg_vai_num = kwargs.get("reg_vai_num", 0)
        rows_num = kwargs.get("rows_num", 0)
        xls_sheet = kwargs.get("reg_type", "pll")

        if xls_sheet == "pll":
            r0 = 0
        else:
            string = ""
            for r0 in range(rows_num):
                buffer1 = df.iloc[r0]
                buffer2 = str(buffer1[0])
                if buffer2.find(string) != -1:
                    break
                else:
                    pass
        reg_arr = []
        for r1 in range(9999):
            buffer1 = df.iloc[r0 + 2 + r1]
            buffer2 = str(buffer1[0])
            if buffer2.find("(Done)") != -1:
                break
            else:
                reg_name = buffer1[0]  # register name
                offset = buffer1[1]  # register offset
                Bit = buffer1[2]  # register S_Bit
                Value = buffer1[reg_vai_num]  # register Value
                TPORT = buffer1[3]  # TPORT Register
                write_en = buffer1[4]  # Read Register
                read_en = buffer1[5]  # Read Register
                Die0 = buffer1[6]  # Die0
                Die1 = buffer1[7]  # Die1
                Die2 = buffer1[8]  # Die2
                V1 = buffer1[9]  # V1
                V2 = buffer1[10]  # V2
                slice0 = buffer1[11]  # slice0
                slice2 = buffer1[12]  # slice2
                reg = f"{offset}, {Bit}, {Value}, {TPORT}, {write_en}, {read_en}, {Die0}, {Die1}, {Die2}, {V1}, {V2}, {slice0}, {slice2}"
                reg_arr += [reg]
                # print(reg_arr)
        return reg_arr

    def seach_xls_reg(self, df, string, **kwargs):
        for r0 in range(9999):
            buffer1 = df.iloc[r0]
            buffer2 = str(buffer1.iloc[0])
            if buffer2.find(string) != -1:
                break
            else:
                pass

        buffer3 = df.iloc[r0]
        self.reg_sequence_loops = int((pd.Series(buffer3)).size)
        for i in range(self.reg_sequence_loops):
            # print(buffer3[i])
            try:
                if buffer3.iloc[i] == (self.register_setup_name) != -1:
                    break
            except:
                pass
        return i

    def seach_xls_reg_lane(self, df, string, **kwargs):
        for r0 in range(9999):
            buffer1 = df.iloc[r0]
            buffer2 = str(buffer1.iloc[0])
            if buffer2.find(string) != -1:
                break
            else:
                pass

        buffer3 = df.iloc[r0]
        self.reg_sequence_loops = int((pd.Series(buffer3)).size)
        for i in range(self.reg_sequence_loops):
            # print(buffer3[i])
            try:
                if buffer3.iloc[i] == (self.even_6) != -1:
                    break
            except:
                pass
        return i

    def write_log(self, content):
        textfile = open(self.save_log, "a+")
        textfile.write(content)
        textfile.close()

    def ChkLog_fail(self, **kwargs):
        find = kwargs.get("find", "NA")

        print("\n< Test Result Summary >")
        import re

        # print(f'Program Interface : {self.prog_HW}')
        buffer = self.m_richText1.Value
        Test_Log = re.sub("\n", "", buffer)
        if Test_Log.find(find) != -1 or Test_Log.find("Failed") != -1:
            self.pass_fail = "FAIL"
            color = "\033"
        else:
            self.pass_fail = "PASS"
            color = "\034"
        print(f"{color}{self.TestItem} : {self.pass_fail}", flush=True)
        return color

    def ChkLog_abp(self, **kwargs):
        find = kwargs.get("find", "NA")

        print("\n< Test Result Summary >")
        import re

        # print(f'Program Interface : {self.prog_HW}')
        buffer = self.m_richText1.Value
        Test_Log = re.sub("\n", "", buffer)
        if Test_Log.find(find) != -1:
            self.abp_pass_fail = "failed"
            color = "\033"
        else:
            self.abp_pass_fail = "PASS"
            color = "\034"
        print(f"{color} EHSOT ABP Enable : {self.abp_pass_fail}", flush=True)
        return color

    def event_window(self):
        def close_window():
            command = "taskkill /f /t /im EXCEL.exe"
            os.system(command)
            window.destroy()

        window = tk.Tk()
        window.title("Eye Scan")
        window.geometry("220x100+0+0")
        window.resizable(False, False)
        window.eval("tk::PlaceWindow . center")
        button = tk.Button(
            text="Press Buttom \n( Close Excel )",
            command=close_window,
            font=("Lucida Bright", 10, "bold"),
            padx=10,
            pady=10,
            bg="#f90",
            bd=5,
        )

        from tkinter.constants import CENTER  # 加到第一行

        button.place(x=110, y=50, anchor=CENTER)

        window.mainloop()

    def clear_text(self, event):
        self.m_richText1.Clear()

    def data_training_event(self, event):
        if self.data_training_en.GetValue() == True:
            self.data_training_en.Label = "Enable"
            self.data_training_en.Value = 1
        else:
            self.data_training_en.Label = "Disable"
            self.data_training_en.Value = 0

    def get_win(self):
        windows = pyautogui.getAllWindows()

        for window in windows:
            buffer = (str(window)).split("title=")[1]
            if buffer.find("Scan") != -1:
                seach = True
                break
            else:
                seach = False
        return seach

    def zero_seach(self, zero_seach):
        if zero_seach[:1] == "1":
            for i in range(99):
                if zero_seach[i : i + 1] == "0":
                    break
                else:
                    pass
            for o in range(99):
                if zero_seach[o + i : o + i + 1] == "1":
                    break
                else:
                    pass
            zero_num = o
        else:
            for i in range(99):
                if zero_seach[i : i + 1] == "1":
                    break
                else:
                    pass
            for o in range(99):
                if zero_seach[o + i : o + i + 1] == "0":
                    break
                else:
                    pass
            for p in range(99):
                if zero_seach[o + p + i : o + p + i + 1] == "1":
                    break
                else:
                    pass
            zero_num = p
        return zero_num


if __name__ == "__main__":
    app = wx.App(False)  # define an object of Application class
    dut = MainFrame(None)  # create gui
    dut.Show(True)  # activate the frame window
    app.MainLoop()  # start the applications
